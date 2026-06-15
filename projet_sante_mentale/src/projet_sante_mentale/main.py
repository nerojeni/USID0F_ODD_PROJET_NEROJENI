"""
main.py
-------
Point d'entrée du projet Santé Mentale des Adolescents.

Usage :
    python main.py                 → pipeline complet
    python main.py --step data     → chargement + CALC seulement
    python main.py --step sheets   → feuilles TCD + TDB seulement
    python main.py --step charts   → graphiques seulement
"""

import argparse
import time
from openpyxl import load_workbook

from projet_sante_mentale.config.settings import PATH_FILE

from projet_sante_mentale.data.loader import (
    load_data,
    build_len_dict,
    init_workbook,
    build_calc_sheet,
    build_correlations_sheet,
)

from projet_sante_mentale.sheets.dashboards import build_tdb1, build_tdb2, build_tdb3
from projet_sante_mentale.sheets.tcd import build_tcd_sheet
from projet_sante_mentale.sheets.tcd2 import build_tcd2_sheet
from projet_sante_mentale.sheets.tcd3 import build_tcd3_sheet

from projet_sante_mentale.charts.charts import (
    build_charts_tdb1,
    build_charts_tdb2,
    build_charts_tdb3,
    build_charts_tdb4,
)


# ─────────────────────────────────────────────────────────────────────────────
# ÉTAPES DU PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def step_data(df=None):
    """
    Étape 1 : Chargement des données et création des feuilles de base.

    Charge le jeu de données, initialise le classeur Excel, génère le dictionnaire 
    des dimensions (len_dict), et construit les feuilles 'DATA', 'CALC' et 'Correlations'.

    Parameters
    ----------
    df : pandas.DataFrame, optional
        Le DataFrame à utiliser. S'il n'est pas fourni, les données seront chargées 
        depuis la source configurée.

    Returns
    -------
    tuple
        Un tuple contenant le DataFrame chargé (df) et le dictionnaire de dimensions (len_dict).
    """
    print("\n" + "=" * 60)
    print("  ÉTAPE 1 — Chargement des données")
    print("=" * 60)

    if df is None:
        df = load_data()

    len_dict = build_len_dict(df)
    init_workbook(df)
    build_calc_sheet(len_dict)
    build_correlations_sheet(df)

    return df, len_dict


def step_sheets(len_dict: dict):
    """
    Étape 2 : Création des feuilles Tableaux de Bord (TDB) et Tableaux Croisés (TCD).

    Génère l'interface des différents tableaux de bord (titres, filtres) et construit 
    les tableaux croisés dynamiques qui stockeront les données calculées.

    Parameters
    ----------
    len_dict : dict
        Dictionnaire contenant les dimensions des variables catégorielles pour 
        ajuster la taille des tableaux.
    """
    print("\n" + "=" * 60)
    print("  ÉTAPE 2 — Création des feuilles")
    print("=" * 60)

    build_tdb1(len_dict)
    build_tdb2(len_dict)
    build_tdb3()
    build_tcd_sheet(len_dict)
    build_tcd2_sheet(len_dict)
    build_tcd3_sheet()


def step_charts(len_dict: dict):
    """
    Étape 3 : Génération et intégration de tous les graphiques.

    Crée les différents graphiques (secteurs, barres, nuages de points, heatmap) 
    à partir des données des TCD et les insère dans les tableaux de bord respectifs.

    Parameters
    ----------
    len_dict : dict
        Dictionnaire contenant les dimensions des variables, nécessaire à la 
        génération de certains graphiques dynamiques.
    """
    print("\n" + "=" * 60)
    print("  ÉTAPE 3 — Création des graphiques")
    print("=" * 60)

    build_charts_tdb1(len_dict)
    build_charts_tdb2(len_dict)
    build_charts_tdb3()
    build_charts_tdb4()


def _print_summary():
    """
    Affiche la liste finale de toutes les feuilles contenues dans le classeur Excel généré.
    Utile pour vérifier le bon déroulement du pipeline.
    """
    wb = load_workbook(PATH_FILE)
    print("\n" + "=" * 60)
    print("  RÉSUMÉ FINAL")
    print("=" * 60)
    print(f"  Fichier : {PATH_FILE}")
    print(f"  Feuilles ({len(wb.sheetnames)}) :")
    for name in wb.sheetnames:
        print(f"      • {name}")
    wb.close()


# ─────────────────────────────────────────────────────────────────────────────
# POINT D'ENTRÉE
# ─────────────────────────────────────────────────────────────────────────────

def main():
    """
    Point d'entrée principal du script. 
    
    Parse les arguments de la ligne de commande pour déterminer quelle(s) étape(s) 
    du pipeline exécuter et chronomètre le temps d'exécution total.
    """
    parser = argparse.ArgumentParser(
        description="Génère le tableau de bord Excel Santé Mentale des Adolescents"
    )
    parser.add_argument(
        "--step",
        choices=["data", "sheets", "charts", "all"],
        default="all",
        help="Étape à exécuter (default: all)",
    )
    args = parser.parse_args()

    t0 = time.time()

    df, len_dict = None, None

    # Exécution conditionnelle des étapes selon l'argument passé
    if args.step in ("data", "all"):
        df, len_dict = step_data()

    if args.step in ("sheets", "all"):
        if len_dict is None:
            # Re-calculer len_dict depuis les données existantes si l'étape data a été ignorée
            df = load_data()
            len_dict = build_len_dict(df)
        step_sheets(len_dict)

    if args.step in ("charts", "all"):
        if len_dict is None:
            # Re-calculer len_dict depuis les données existantes si l'étape sheets a été exécutée seule
            df = load_data()
            len_dict = build_len_dict(df)
        step_charts(len_dict)

    _print_summary()

    elapsed = time.time() - t0
    print(f"\n  ⏱  Terminé en {elapsed:.1f}s\n")


if __name__ == "__main__":
    main()