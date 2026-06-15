"""
utils/excel_helpers.py
-----------------------
Fonctions utilitaires génériques pour la manipulation de feuilles openpyxl.
"""

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from projet_sante_mentale.utils.styles import (
    TITLE_FILL,
    TITLE_FONT_SM,
    HEADER_FILL,
    HEADER_FONT,
    ROW_HEADER_FILL,
    ROW_HEADER_FONT,
    BORDER,
    CENTER,
    LEFT,
)
from projet_sante_mentale.config.settings import COLORS


# ─── ÉCRITURE DE TABLEAUX ─────────────────────────────────────────────────────
def write_table(
    ws, df, start_row: int, title: str = None, space: int = 3, padding: int = 2
) -> int:
    """
    Écrit un DataFrame dans une feuille Excel à partir de start_row.

    Retourne la première ligne disponible après le tableau.
    """
    if title:
        ws.cell(row=start_row, column=1, value=title)
        start_row += 1

    for col_idx, header in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=col_idx, value=header)
        col_letter = get_column_letter(col_idx)
        width = len(str(header)) + padding
        current = ws.column_dimensions[col_letter].width
        if current is None or width > current:
            ws.column_dimensions[col_letter].width = width

    for i, row in df.iterrows():
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=start_row + i + 1, column=col_idx, value=value)

    return start_row + len(df) + space + 1


# ─── TITRES DE SECTION ────────────────────────────────────────────────────────
def write_section_title(
    ws, row: int, text: str, merge_range: str = None, height: int = 25
) -> None:
    """
    Écrit un titre de section avec fond bleu foncé + police blanche.
    merge_range : ex. 'A1:J1'
    """
    first_col = merge_range.split(":")[0] if merge_range else f"A{row}"
    first_cell = first_col if ":" not in first_col else merge_range.split(":")[0]
    cell = ws[f"{first_cell[0]}{row}"] if len(first_cell) == 1 else ws[first_cell]

    ws.cell(row=row, column=1, value=text)
    ws[f"A{row}"].font = TITLE_FONT_SM
    ws[f"A{row}"].fill = TITLE_FILL
    ws[f"A{row}"].alignment = CENTER

    if merge_range:
        ws.merge_cells(merge_range)
    ws.row_dimensions[row].height = height


def write_col_headers(
    ws, row: int, headers: list, start_col: int = 1, widths: list = None
) -> None:
    """Écrit une ligne d'en-têtes avec style teal."""
    for i, h in enumerate(headers):
        col = start_col + i
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = CENTER
        if widths and i < len(widths):
            ws.column_dimensions[get_column_letter(col)].width = widths[i]


def write_row_header(ws, row: int, col: int, value) -> None:
    """Écrit un en-tête de ligne avec style teal clair."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = ROW_HEADER_FILL
    cell.font = ROW_HEADER_FONT
    cell.border = BORDER
    cell.alignment = CENTER


# ─── FILTRES INTERACTIFS ──────────────────────────────────────────────────────
def add_filter(
    worksheet,
    start_col: str,
    filter_row: int,
    title_text: str,
    data_source_col: str,
    len_data: int,
    helper_col: str,
    default_value: str = "Tous",
    title_color: str = "FF003D6B",
    value_color: str = "FFFFFF",
) -> None:
    """
    Crée un filtre horizontal (titre + cellule de valeur + DataValidation).

    Parameters
    ----------
    worksheet      : feuille cible
    start_col      : colonne du titre (ex. 'C')
    filter_row     : ligne du filtre (ex. 3)
    title_text     : texte du titre (ex. 'Âge')
    data_source_col: colonne CALC source (ex. 'C')
    len_data       : nombre de valeurs uniques (issu de len_dict)
    helper_col     : colonne cachée helper (ex. 'AA')
    default_value  : valeur par défaut ('Tous')
    title_color    : fond hexadécimal du titre
    value_color    : fond hexadécimal de la valeur
    """
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

    value_col = chr(ord(start_col) + 1)

    t_fill = PatternFill(
        start_color=title_color, end_color=title_color, fill_type="solid"
    )
    v_fill = PatternFill(
        start_color=value_color, end_color=value_color, fill_type="solid"
    )
    align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Titre
    tc = worksheet[f"{start_col}{filter_row}"]
    tc.value, tc.fill, tc.font = (
        title_text,
        t_fill,
        Font(bold=True, color="FFFFFF", size=10),
    )
    tc.alignment, tc.border = align, border
    worksheet.column_dimensions[start_col].width = 12

    # Valeur (cellule éditable)
    vc = worksheet[f"{value_col}{filter_row}"]
    vc.value, vc.fill, vc.font = (
        default_value,
        v_fill,
        Font(bold=True, color="000000", size=10),
    )
    vc.alignment, vc.border = align, border
    worksheet.column_dimensions[value_col].width = 12

    # Colonne helper (cachée)
    worksheet[f"{helper_col}1"] = "Tous"
    for i in range(2, len_data + 1):
        worksheet[f"{helper_col}{i}"] = f"=CALC!{data_source_col}{i}"
    worksheet.column_dimensions[helper_col].hidden = True

    # DataValidation dropdown
    formula = f"=${helper_col}$1:${helper_col}${len_data}"
    dv = DataValidation(type="list", formula1=formula, allow_blank=False)
    dv.error = "Sélectionnez une valeur valide"
    dv.errorTitle = "Entrée invalide"
    dv.prompt = f"Sélectionnez un {title_text.lower()}"
    dv.promptTitle = "Filtrer"
    worksheet.add_data_validation(dv)
    dv.add(f"{value_col}{filter_row}")

    print(
        f"   ✅ Filtre '{title_text}' → {start_col}{filter_row}:{value_col}{filter_row}"
    )


# ─── MISE EN PAGE D'UNE PAGE DE TABLEAU DE BORD ──────────────────────────────
def setup_dashboard_page(
    ws, title_text: str, merge_cols: str = "A1:O1", row_height: int = 30
) -> None:
    """
    Initialise une page de tableau de bord :
    - Grille masquée
    - Titre principal bleu foncé centré
    """
    ws.sheet_view.showGridLines = False

    ws.merge_cells(merge_cols)
    ws["A1"].value = title_text
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=COLORS["white"])
    ws["A1"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.row_dimensions[1].height = row_height


def setup_filter_row(ws, filter_row: int, len_dict: dict, tdb_prefix: str) -> None:
    """
    Écrit la ligne 'Filtres :' et les 4 filtres horizontaux.

    tdb_prefix : 'TDB1' ou 'TDB2' → détermine les colonnes helper (AA-AD ou AE-AH)
    """
    helpers = {"TDB1": ["AA", "AB", "AC", "AD"], "TDB2": ["AE", "AF", "AG", "AH"]}[
        tdb_prefix
    ]

    ws[f"A{filter_row}"] = "Filtres :"
    ws[f"A{filter_row}"].font = Font(size=10, bold=True)
    ws[f"A{filter_row}"].alignment = LEFT
    ws.row_dimensions[filter_row].height = 25

    add_filter(
        ws,
        "C",
        filter_row,
        "Âge",
        "C",
        len_dict["len_age"],
        helpers[0],
        title_color="FF003D6B",
        value_color="FFFFFF",
    )
    add_filter(
        ws,
        "F",
        filter_row,
        "Genre",
        "A",
        len_dict["len_gender"],
        helpers[1],
        title_color="FF003D6B",
        value_color="FFFFFF",
    )
    add_filter(
        ws,
        "I",
        filter_row,
        "Plateforme",
        "E",
        len_dict["len_platform_usage"],
        helpers[2],
        title_color="FF003D6B",
        value_color="FFFFFF",
    )
    add_filter(
        ws,
        "L",
        filter_row,
        "Interaction",
        "G",
        len_dict["len_social_interaction_level"],
        helpers[3],
        title_color="FF003D6B",
        value_color="FFFFFF",
    )
