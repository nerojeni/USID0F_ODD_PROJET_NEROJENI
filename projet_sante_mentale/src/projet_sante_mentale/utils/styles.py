"""
utils/styles.py
---------------
Styles openpyxl centralisés, réutilisables dans tous les modules de génération Excel.
"""

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from projet_sante_mentale.config.settings import COLORS


def make_fill(hex_color: str) -> PatternFill:
    """
    Crée un remplissage de cellule uni (solid) avec la couleur hexadécimale spécifiée.
    """
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def make_border(style: str = "thin") -> Border:
    """
    Crée une bordure standard uniforme (fine par défaut) pour les 4 côtés d'une cellule.
    """
    side = Side(style=style)
    return Border(left=side, right=side, top=side, bottom=side)


def make_border_colored(color: str = "DDDDDD") -> Border:
    """
    Crée une bordure fine avec une couleur personnalisée (gris clair par défaut).
    """
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


# ─── Objets de style pré-construits ───────────────────────────────────────────

# Alignements
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")

# Bordures
BORDER = make_border()
BORDER_LIGHT = make_border_colored()

# Titres principaux (bandeau bleu foncé)
TITLE_FILL = make_fill(COLORS["title_bg"])
TITLE_FONT = Font(bold=True, size=14, color=COLORS["white"])
TITLE_FONT_SM = Font(bold=True, size=12, color=COLORS["white"])

# En-têtes de colonnes (teal standard)
HEADER_FILL = make_fill(COLORS["header_bg"])
HEADER_FONT = Font(bold=True, size=10, color=COLORS["white"])

# En-têtes de lignes / lignes paires (teal clair)
ROW_HEADER_FILL = make_fill(COLORS["row_header_bg"])
ROW_HEADER_FONT = Font(bold=True, size=10, color=COLORS["black"])

# Styles dédiés aux blocs d'Indicateurs Clés de Performance (KPI)
KPI_TITLE_FILL = make_fill(COLORS["title_bg"])
KPI_TITLE_FONT = Font(bold=True, size=11, color=COLORS["white"])
KPI_VALUE_FILL = make_fill(COLORS["white"])
KPI_VALUE_FONT = Font(bold=True, size=14, color=COLORS["title_bg"])
KPI_VALUE_FONT_LG = Font(bold=True, size=16, color=COLORS["white"])

# Couleurs par genre (ex: graphiques sectoriels ou KPI genrés)
KPI_MALE_FILL = make_fill(COLORS["male"])
KPI_FEMALE_FILL = make_fill(COLORS["female"])