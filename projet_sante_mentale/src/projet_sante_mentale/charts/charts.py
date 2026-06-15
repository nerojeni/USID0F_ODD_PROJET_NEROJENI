"""
charts/charts.py
----------------
Création de tous les graphiques (BarChart, PieChart, ScatterChart).
"""

from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference, Series
from openpyxl.chart import ScatterChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font

from projet_sante_mentale.config.settings import PATH_FILE, COLORS


def build_all_charts(len_dict: dict) -> None:
    """
    Point d'entrée unique : crée tous les graphiques dans TDB1, TDB2 et TDB3.

    Appelle successivement build_charts_tdb1, build_charts_tdb2 et
    build_charts_tdb3 dans l'ordre d'affichage du tableau de bord.

    Parameters
    ----------
    len_dict : dict
        Dictionnaire des longueurs de valeurs uniques par colonne,
        produit par build_len_dict() (ex. {"len_age": 9, "len_gender": 3, ...}).
    """
    build_charts_tdb1(len_dict)
    build_charts_tdb2(len_dict)
    build_charts_tdb3()


# ─────────────────────────────────────────────────────────────────────────────
# TDB1 : BarChart Addiction + PieChart Dépression + Boxplot Addiction
# ─────────────────────────────────────────────────────────────────────────────


def build_charts_tdb1(len_dict: dict) -> None:
    """
    Crée et place les trois graphiques de la page TDB1, ainsi que les cellules KPI.

    Graphiques ajoutés :
      - B5  : BarChart addiction moyenne par âge et genre.
      - J5  : PieChart répartition dépression selon le genre.
      - B16 : BarChart boxplot addiction selon l'âge.

    Ouvre le fichier Excel, insère les graphiques dans la feuille TDB1
    en lisant les données de la feuille TCD, puis sauvegarde.

    Parameters
    ----------
    len_dict : dict
        Dictionnaire des longueurs de valeurs uniques (voir build_len_dict).
    """
    wb = load_workbook(PATH_FILE)
    ws_tcd = wb["TCD"]
    ws_tdb1 = wb["TDB1"]

    ws_tdb1.add_chart(_barchart_addiction(ws_tcd, len_dict), "B5")
    ws_tdb1.add_chart(_piechart_depression(ws_tcd), "J5")
    ws_tdb1.add_chart(_barchart_boxplot(ws_tcd, len_dict), "B16")
    _write_kpi_cells_tdb1(ws_tcd, ws_tdb1)

    wb.save(PATH_FILE)
    wb.close()
    print("Graphiques TDB1 créés (BarChart addiction + PieChart + Boxplot + KPI)")


def _barchart_addiction(ws_tcd, len_dict: dict) -> BarChart:
    """
    Construit un BarChart groupé de l'addiction moyenne par âge et genre.

    Les données sont lues depuis les lignes 11-13 de la feuille TCD :
      - Ligne 11 : en-têtes des âges (catégories de l'axe X)
      - Lignes 12-13 : moyennes d'addiction pour Male et Female

    Les étiquettes de valeurs sont affichées en position 'outEnd'.
    Les couleurs sont appliquées via _apply_series_colors.

    Parameters
    ----------
    ws_tcd : Worksheet
        Feuille TCD contenant les données calculées.
    len_dict : dict
        Dictionnaire des longueurs ; ``len_age`` détermine le nombre de colonnes.

    Returns
    -------
    BarChart
        Graphique prêt à être ajouté à une feuille via add_chart().
    """
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Addiction moyenne par âge et genre"
    chart.y_axis.title = "Niveau d'addiction"
    chart.x_axis.title = "Âge"
    chart.height = 5
    chart.width = 15

    max_col = len_dict["len_age"]
    data = Reference(ws_tcd, min_col=1, min_row=12, max_col=max_col, max_row=13)
    cats = Reference(ws_tcd, min_col=2, min_row=11, max_col=max_col)

    chart.add_data(data, titles_from_data=True, from_rows=True)
    chart.set_categories(cats)
    chart.overlap = -15

    for series in chart.series:
        series.dLbls = DataLabelList()
        series.dLbls.showVal = True
        series.dLbls.showCatName = False
        series.dLbls.showSerName = False
        series.dLbls.position = "outEnd"
        series.dLbls.numFmt = "0"

    _apply_series_colors(chart, [COLORS["male"], COLORS["female"]])
    chart.legend.position = "r"
    chart.y_axis.majorGridlines = None
    return chart


def _piechart_depression(ws_tcd) -> PieChart:
    """
    Construit un PieChart de la répartition de la dépression selon le genre.

    Les données sont lues depuis le bloc helper situé à partir de la ligne 120
    de la feuille TCD (Male en ligne 122, Female en ligne 123).
    Les pourcentages sont affichés à la place des valeurs brutes.
    La légende est supprimée au profit des noms de catégories sur le graphique.

    Parameters
    ----------
    ws_tcd : Worksheet
        Feuille TCD contenant les helpers de dépression par genre.

    Returns
    -------
    PieChart
        Graphique prêt à être ajouté à une feuille via add_chart().
    """
    start = 120
    chart = PieChart()
    chart.title = "Répartition de la dépression selon le genre"
    chart.height = 5
    chart.width = 10

    data_ref = Reference(ws_tcd, min_col=2, min_row=start + 1, max_row=start + 3)
    cats_ref = Reference(ws_tcd, min_col=1, min_row=start + 2, max_row=start + 3)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    chart.dLbls = DataLabelList()
    chart.dLbls.showPercent = True
    chart.dLbls.showVal = False
    chart.dLbls.showCatName = True
    chart.legend = None

    series = chart.series[0]
    for idx, color in enumerate([COLORS["male"], COLORS["female"]]):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        series.dPt.append(pt)

    return chart


def _barchart_boxplot(ws_tcd, len_dict: dict) -> BarChart:
    """
    Construit un BarChart empilé simulant un boxplot de l'addiction selon l'âge.

    Les 5 séries (Min, Q1, Médiane, Q3, Max) sont lues depuis les colonnes B-F
    du tableau statistique de la feuille TCD (à partir de la ligne 21).
    Chaque série a une couleur distincte pour imiter visuellement un boxplot.

    Parameters
    ----------
    ws_tcd : Worksheet
        Feuille TCD contenant les statistiques d'addiction par âge.
    len_dict : dict
        Dictionnaire des longueurs ; ``len_age`` détermine le nombre de lignes.

    Returns
    -------
    BarChart
        Graphique prêt à être ajouté à une feuille via add_chart().
    """
    start = 21
    num_ages = len_dict["len_age"] - 1

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Répartition niveau Addiction selon l'âge"
    chart.y_axis.title = "Niveau d'addiction"
    chart.x_axis.title = "Âge"
    chart.height = 7
    chart.width = 15

    cats_ref = Reference(ws_tcd, min_col=1, min_row=start + 1, max_row=start + num_ages)

    series_config = [
        ("Min", "808080"),
        ("Q1", "A6A6A6"),
        ("Médiane", "FF1493"),
        ("Q3", "CCCCCC"),
        ("Max", "555555"),
    ]

    for col_idx, (name, color) in enumerate(series_config):
        data = Reference(
            ws_tcd, min_col=col_idx + 2, min_row=start, max_row=start + num_ages
        )
        chart.add_data(data, titles_from_data=True)
        chart.series[-1].graphicalProperties.solidFill = color

        lbl = DataLabelList()
        lbl.showVal = True
        lbl.position = "outEnd"
        lbl.numFmt = "0"
        chart.series[-1].dLbls = lbl

    chart.set_categories(cats_ref)
    chart.overlap = -50
    chart.legend.position = "r"
    chart.y_axis.majorGridlines = None
    return chart


def _write_kpi_cells_tdb1(ws_tcd, ws_tdb1) -> None:
    """
    Écrit les cellules KPI dans la page TDB1 à partir de la colonne J, ligne 16.

    Quatre indicateurs sont affichés en deux colonnes (J-K et L-M) :
      - J16:K16 / J17:K17 : Réseau populaire (source TCD!A32)
      - L16:M16 / L17:M17 : Nombre de dépressifs (source TCD!B34)
      - J19:K19 / J20:K20 : Performance scolaire moyenne (source TCD!D33)
      - L19:M19 / L20:M20 : Min et Max performance (sources TCD!E34 et TCD!F34)

    Chaque KPI est composé d'une cellule titre (fond bleu foncé, texte blanc)
    et d'une cellule valeur (fond blanc, texte bleu foncé, police plus grande).
    Les cellules Min et Max sont sur la même ligne sans fusion.

    Parameters
    ----------
    ws_tcd : Worksheet
        Feuille TCD source des valeurs KPI (non modifiée par cette fonction).
    ws_tdb1 : Worksheet
        Feuille TDB1 cible où les cellules KPI sont écrites.
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    kpi_row = 16
    FILL_DARK = PatternFill(
        start_color=COLORS["title_bg"], end_color=COLORS["title_bg"], fill_type="solid"
    )
    FILL_LIGHT = PatternFill(
        start_color=COLORS["white"], end_color=COLORS["white"], fill_type="solid"
    )
    FONT_TITLE = Font(bold=True, color=COLORS["white"], size=10)
    FONT_VAL = Font(bold=True, size=14, color=COLORS["title_bg"])
    FONT_SMALL = Font(bold=True, size=11, color=COLORS["title_bg"])
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _kpi_title(cell_ref, text, merge_to=None):
        """Écrit une cellule titre KPI avec fond bleu foncé et texte blanc."""
        c = ws_tdb1[cell_ref]
        c.value, c.fill, c.font, c.alignment, c.border = (
            text,
            FILL_DARK,
            FONT_TITLE,
            CENTER,
            BORDER,
        )
        if merge_to:
            ws_tdb1.merge_cells(f"{cell_ref}:{merge_to}")

    def _kpi_value(cell_ref, formula, merge_to=None, fmt="@", large=True):
        """
        Écrit une cellule valeur KPI avec fond blanc et texte bleu foncé.

        Parameters
        ----------
        cell_ref : str  Référence de la cellule (ex. 'J17').
        formula  : str  Formule Excel ou valeur à écrire.
        merge_to : str  Cellule de fin de fusion (optionnel).
        fmt      : str  Format numérique Excel (ex. '0.00', '0', '@').
        large    : bool Si True, police taille 14 ; sinon taille 11.
        """
        c = ws_tdb1[cell_ref]
        c.value, c.fill = formula, FILL_LIGHT
        c.font = Font(bold=True, size=14 if large else 11, color=COLORS["title_bg"])
        c.alignment, c.border, c.number_format = CENTER, BORDER, fmt
        ws_tdb1.row_dimensions[int(cell_ref[1:])].height = 30
        if merge_to:
            ws_tdb1.merge_cells(f"{cell_ref}:{merge_to}")

    _kpi_title(f"J{kpi_row}", "Réseau populaire", f"K{kpi_row}")
    _kpi_value(f"J{kpi_row + 1}", "=TCD!A32", f"K{kpi_row + 1}")

    _kpi_title(f"L{kpi_row}", "Dépressifs", f"M{kpi_row}")
    _kpi_value(f"L{kpi_row + 1}", "=TCD!B34", f"M{kpi_row + 1}", fmt="0")

    _kpi_title(f"J{kpi_row + 3}", "Performance scolaire", f"K{kpi_row + 3}")
    _kpi_value(f"J{kpi_row + 4}", "=TCD!D33", f"K{kpi_row + 4}", fmt="0.00")

    _kpi_title(f"L{kpi_row + 3}", "Min / Max", f"M{kpi_row + 3}")

    for cell_ref, formula in [
        (f"L{kpi_row + 4}", "=TCD!E34"),
        (f"M{kpi_row + 4}", "=TCD!F34"),
    ]:
        ws_tdb1[cell_ref].value = formula
        ws_tdb1[cell_ref].fill = FILL_LIGHT
        ws_tdb1[cell_ref].font = FONT_SMALL
        ws_tdb1[cell_ref].alignment = CENTER
        ws_tdb1[cell_ref].border = BORDER
        ws_tdb1[cell_ref].number_format = "0.00"

    for col in ["J", "K", "L", "M"]:
        ws_tdb1.column_dimensions[col].width = 14


# ─────────────────────────────────────────────────────────────────────────────
# TDB2 : KPI Genres + BarChart Perf Scolaire + BarChart Dépressifs par Âge
# ─────────────────────────────────────────────────────────────────────────────


def build_charts_tdb2(len_dict: dict) -> None:
    """
    Crée et place les graphiques et KPI de la page TDB2.

    Éléments ajoutés :
      - B6:E11 : KPI dépressifs + addiction par genre (cellules colorées Male/Female).
      - G5     : BarChart performance scolaire selon le temps d'écran.
      - L5     : BarChart horizontal dépressifs par âge.

    Ouvre le fichier Excel, écrit dans TDB2 en lisant TCD et TCD2, puis sauvegarde.

    Parameters
    ----------
    len_dict : dict
        Dictionnaire des longueurs de valeurs uniques (voir build_len_dict).
    """
    wb = load_workbook(PATH_FILE)
    ws_tcd2 = wb["TCD2"]
    ws_tcd = wb["TCD"]
    ws_tdb2 = wb["TDB2"]

    _write_kpi_tdb2(ws_tcd, ws_tdb2)
    ws_tdb2.add_chart(_barchart_perf_scolaire(ws_tcd2), "G5")
    ws_tdb2.add_chart(_barchart_depressifs_age(ws_tcd2, len_dict), "L5")

    wb.save(PATH_FILE)
    wb.close()
    print("Graphiques TDB2 créés (KPI + BarChart perf + BarChart dépressifs)")


def _write_kpi_tdb2(ws_tcd, ws_tdb2) -> None:
    """
    Écrit deux blocs KPI dans la page TDB2 à partir de la ligne 6.

    Bloc 1 (ligne 6-7) : Nombre de dépressifs par genre
      - B7:C7 → Male (fond bleu foncé #040459), source TCD!B101
      - D7:E7 → Female (fond bordeaux #7D093F), source TCD!B102

    Bloc 2 (ligne 10-11) : Addiction moyenne par genre
      - B11:C11 → Male, source TCD!B106
      - D11:E11 → Female, source TCD!B107

    Chaque bloc a un titre fusionné B-E en fond bleu foncé et deux cellules
    valeur côte à côte colorées selon le genre.

    Parameters
    ----------
    ws_tcd : Worksheet
        Feuille TCD source des helpers (non modifiée ici).
    ws_tdb2 : Worksheet
        Feuille TDB2 cible où les cellules KPI sont écrites.
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    FILL_DARK = PatternFill(
        start_color=COLORS["title_bg"], end_color=COLORS["title_bg"], fill_type="solid"
    )
    FILL_MALE = PatternFill(
        start_color=COLORS["male"], end_color=COLORS["male"], fill_type="solid"
    )
    FILL_FEMALE = PatternFill(
        start_color=COLORS["female"], end_color=COLORS["female"], fill_type="solid"
    )
    FONT_TITLE = Font(bold=True, color=COLORS["white"], size=11)
    FONT_VAL = Font(bold=True, size=16, color=COLORS["white"])
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    CENTER = Alignment(horizontal="center", vertical="center")

    kpi_configs = [
        (6, "Nombre de Dépressifs selon le genre", "=TCD!B101", "=TCD!B102", "0"),
        (10, "Niveau Addiction moyen selon le genre", "=TCD!B106", "=TCD!B107", "0.0"),
    ]

    for kpi_row, title, male_formula, female_formula, fmt in kpi_configs:
        ws_tdb2[f"B{kpi_row}"] = title
        ws_tdb2[f"B{kpi_row}"].fill = FILL_DARK
        ws_tdb2[f"B{kpi_row}"].font = FONT_TITLE
        ws_tdb2[f"B{kpi_row}"].alignment = CENTER
        ws_tdb2[f"B{kpi_row}"].border = BORDER
        ws_tdb2.merge_cells(f"B{kpi_row}:E{kpi_row}")
        ws_tdb2.row_dimensions[kpi_row].height = 20

        val_row = kpi_row + 1
        ws_tdb2.row_dimensions[val_row].height = 35

        for col_start, col_end, formula, fill in [
            ("B", "C", male_formula, FILL_MALE),
            ("D", "E", female_formula, FILL_FEMALE),
        ]:
            ws_tdb2[f"{col_start}{val_row}"] = formula
            ws_tdb2[f"{col_start}{val_row}"].fill = fill
            ws_tdb2[f"{col_start}{val_row}"].font = FONT_VAL
            ws_tdb2[f"{col_start}{val_row}"].alignment = CENTER
            ws_tdb2[f"{col_start}{val_row}"].border = BORDER
            ws_tdb2[f"{col_start}{val_row}"].number_format = fmt
            ws_tdb2.merge_cells(f"{col_start}{val_row}:{col_end}{val_row}")


def _barchart_perf_scolaire(ws_tcd2) -> BarChart:
    """
    Construit un BarChart groupé de la performance scolaire selon le temps d'écran.

    Les données sont lues depuis le Tableau 1 de TCD2 :
      - Ligne 2       : en-têtes des catégories de temps d'écran (0-1h, 1-2h, 2-3h, 3-4h)
      - Lignes 3 et 4 : moyennes de performance pour Male et Female

    from_rows=True est utilisé car chaque ligne correspond à une série (Male/Female).
    Les étiquettes de valeurs sont affichées avec 2 décimales.

    Parameters
    ----------
    ws_tcd2 : Worksheet
        Feuille TCD2 contenant le tableau de performance scolaire.

    Returns
    -------
    BarChart
        Graphique prêt à être ajouté à une feuille via add_chart().
    """
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Performance scolaire moyenne par temps d'écran"
    chart.y_axis.title = "Performance scolaire"
    chart.x_axis.title = "Temps d'écran (Heures)"
    chart.height = 7
    chart.width = 10

    data = Reference(ws_tcd2, min_col=1, min_row=3, max_col=5, max_row=4)
    cats = Reference(ws_tcd2, min_col=2, min_row=2, max_col=5)

    chart.add_data(data, titles_from_data=True, from_rows=True)
    chart.set_categories(cats)
    chart.overlap = -15

    for series in chart.series:
        series.dLbls = DataLabelList()
        series.dLbls.showVal = True
        series.dLbls.position = "outEnd"
        series.dLbls.numFmt = "0.00"

    _apply_series_colors(chart, [COLORS["male"], COLORS["female"]])
    chart.legend.position = "r"
    chart.y_axis.majorGridlines = None
    return chart


def _barchart_depressifs_age(ws_tcd2, len_dict: dict) -> BarChart:
    """
    Construit un BarChart horizontal du nombre de dépressifs par âge.

    Les données sont lues depuis le Tableau 2 de TCD2 (à partir de la ligne 10) :
      - Colonne A : âges (catégories de l'axe Y)
      - Colonne B : nombre de dépressifs par âge

    Le graphique est orienté horizontalement (type="bar") pour une meilleure
    lisibilité des étiquettes d'âge. La couleur bordeaux (#7D093F) rappelle
    visuellement l'association à la dépression.

    Parameters
    ----------
    ws_tcd2 : Worksheet
        Feuille TCD2 contenant le tableau des dépressifs par âge.
    len_dict : dict
        Dictionnaire des longueurs ; ``len_age`` détermine la dernière ligne de données.

    Returns
    -------
    BarChart
        Graphique horizontal prêt à être ajouté à une feuille via add_chart().
    """
    start = 10
    header_row = start + 1
    last_row = header_row + len_dict["len_age"] - 1

    chart = BarChart()
    chart.type = "bar"  # Horizontal
    chart.title = "Nombre de dépressifs par Âge"
    chart.height = 7
    chart.width = 7

    data = Reference(ws_tcd2, min_col=2, min_row=header_row, max_row=last_row)
    cats = Reference(ws_tcd2, min_col=1, min_row=header_row + 1, max_row=last_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.position = "outEnd"
    chart.series[0].graphicalProperties.solidFill = COLORS["female"]

    chart.legend.position = "r"
    chart.y_axis.majorGridlines = None
    return chart


# ─────────────────────────────────────────────────────────────────────────────
# TDB3 : 4 Scatter Plots
# ─────────────────────────────────────────────────────────────────────────────


def build_charts_tdb3() -> None:
    """
    Crée et place les quatre scatter plots de la page TDB3.

    Les graphiques sont configurés via une liste de dictionnaires (charts_config)
    et construits par la fonction _make_scatter. Chaque scatter plot lit ses
    données dans la feuille TCD3 selon les plages de lignes définies dans config.

    Positions dans TDB3 :
      - B4  : Sommeil vs Addiction par Interaction Sociale
      - K4  : Performance Scolaire selon Stress par Genre
      - B24 : Anxiété selon Temps d'Écran par Interaction sociale
      - K24 : Performance Scolaire selon Sommeil par État Dépressif

    Ouvre le fichier Excel, insère les graphiques dans TDB3, puis sauvegarde.
    """
    wb = load_workbook(PATH_FILE)
    ws_tcd3 = wb["TCD3"]
    ws_tdb3 = wb["TDB3"]

    charts_config = [
        {
            "title": "Sommeil vs Addiction par Interaction Sociale",
            "x_title": "Temps de Sommeil (heures)",
            "y_title": "Niveau d'Addiction",
            "min_row": 3,
            "max_row": 12,
            "header_row": 2,
            "labels": ["Interaction faible", "Interaction élevée"],
            "colors": [COLORS["scatter_teal"], COLORS["scatter_rose"]],
            "position": "B4",
        },
        {
            "title": "Performance Scolaire selon Stress par Genre",
            "x_title": "Niveau de Stress",
            "y_title": "Performance Scolaire",
            "min_row": 17,
            "max_row": 26,
            "header_row": 16,
            "labels": ["Performance (Male)", "Performance (Female)"],
            "colors": [COLORS["scatter_blue"], COLORS["scatter_pink"]],
            "position": "K4",
        },
        {
            "title": "Anxiété selon Temps d'Écran par Interaction sociale",
            "x_title": "Temps d'Écran avant Sommeil (heures)",
            "y_title": "Niveau d'Anxiété",
            "min_row": 32,
            "max_row": 38,
            "header_row": 31,
            "labels": ["Anxiété (Low)", "Anxiété (High)"],
            "colors": [COLORS["scatter_teal"], COLORS["scatter_rose"]],
            "position": "B24",
        },
        {
            "title": "Performance Scolaire selon Sommeil par État Dépressif",
            "x_title": "Temps de Sommeil (heures)",
            "y_title": "Performance Scolaire",
            "min_row": 44,
            "max_row": 53,
            "header_row": 43,
            "labels": ["Non dépressive", "Dépressive"],
            "colors": [COLORS["scatter_teal"], COLORS["scatter_rose"]],
            "position": "K24",
        },
    ]

    for cfg in charts_config:
        chart = _make_scatter(
            ws_tcd3, **{k: v for k, v in cfg.items() if k != "position"}
        )
        ws_tdb3.add_chart(chart, cfg["position"])

    wb.save(PATH_FILE)
    wb.close()
    print("Graphiques TDB3 créés (4 scatter plots)")


def _make_scatter(
    ws_tcd3,
    title: str,
    x_title: str,
    y_title: str,
    min_row: int,
    max_row: int,
    header_row: int,
    labels: list,
    colors: list,
    **_,
) -> ScatterChart:
    """
    Construit un ScatterChart (nuage de points) à partir d'un tableau TCD3.

    Structure attendue dans ws_tcd3 :
      - Colonne A (lignes min_row:max_row) : valeurs X numériques
      - Colonne B (lignes header_row:max_row) : valeurs Y de la série 1 (avec en-tête)
      - Colonne C (lignes header_row:max_row) : valeurs Y de la série 2 (avec en-tête)

    Les titres de séries sont lus depuis header_row via title_from_data=True.
    Les lignes de connexion entre les points sont désactivées (noFill=True)
    pour obtenir un nuage de points pur. Les marqueurs circulaires sont colorés
    individuellement pour chaque série.

    Parameters
    ----------
    ws_tcd3 : Worksheet
        Feuille TCD3 contenant les données des 4 tableaux scatter.
    title : str
        Titre du graphique affiché dans Excel.
    x_title : str
        Label de l'axe horizontal.
    y_title : str
        Label de l'axe vertical.
    min_row : int
        Première ligne de données X (sans en-tête).
    max_row : int
        Dernière ligne de données.
    header_row : int
        Ligne contenant les en-têtes de séries Y (title_from_data).
    labels : list of str
        Noms des deux séries (utilisés si title_from_data échoue).
    colors : list of str
        Couleurs hexadécimales sans '#' pour les marqueurs des deux séries.
    **_ :
        Arguments supplémentaires ignorés (ex. 'position' du dict de config).

    Returns
    -------
    ScatterChart
        Graphique prêt à être ajouté à une feuille via add_chart().
    """
    chart = ScatterChart()
    chart.title = title
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.height = 9
    chart.width = 14
    chart.style = 13
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    xvalues = Reference(ws_tcd3, min_col=1, min_row=min_row, max_row=max_row)

    for i, (label, color) in enumerate(zip(labels, colors)):
        col = 2 + i
        yvalues = Reference(ws_tcd3, min_col=col, min_row=header_row, max_row=max_row)
        serie = Series(yvalues, xvalues, title_from_data=True)
        serie.marker.symbol = "circle"
        serie.marker.size = 8
        serie.graphicalProperties.line.noFill = True
        serie.marker.graphicalProperties.solidFill = color
        serie.marker.graphicalProperties.line.solidFill = color
        chart.series.append(serie)

    chart.legend.position = "r"
    chart.y_axis.majorGridlines = None
    return chart


# ─────────────────────────────────────────────────────────────────────────────
# TDB4 : Heatmap de corrélation
# ─────────────────────────────────────────────────────────────────────────────


def build_charts_tdb4() -> None:
    """
    Crée la page TDB4 avec une heatmap de la matrice de corrélation.

    La heatmap est construite en liant dynamiquement les cellules de TDB4
    aux valeurs calculées dans la feuille Correlations (formules =Correlations!Xn).
    Un formatage conditionnel ColorScale à 3 couleurs est ensuite appliqué :
      - Rouge très clair (#FFF2F2) pour les corrélations négatives (-1.0)
      - Blanc (#FFFFFF) pour les corrélations nulles (0.0)
      - Bleu très clair (#E6F2FF) pour les corrélations positives (1.0)

    Les variables analysées sont définies dans COLS_CORR (config/settings.py).
    La feuille TDB4 est recréée à chaque appel (suppression si elle existe déjà).

    Structure de TDB4 :
      - Ligne 1      : Titre principal fusionné
      - Ligne 4      : En-têtes de colonnes (noms de variables)
      - Lignes 5-14  : Matrice 10×10 avec liens vers Correlations + heatmap
      - Colonne A    : En-têtes de lignes (noms de variables)
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    from config.settings import COLS_CORR

    wb = load_workbook(PATH_FILE)

    if "TDB4" in wb.sheetnames:
        del wb["TDB4"]
    TDB4 = wb.create_sheet("TDB4", 3)
    TDB4.sheet_view.showGridLines = False

    FILL_DARK = PatternFill(
        start_color=COLORS["title_bg"], end_color=COLORS["title_bg"], fill_type="solid"
    )
    FILL_HEAD = PatternFill(
        start_color=COLORS["header_bg"],
        end_color=COLORS["header_bg"],
        fill_type="solid",
    )
    FONT_MAIN = Font(name="Calibri", size=14, bold=True, color=COLORS["white"])
    FONT_HEAD = Font(bold=True, size=10, color=COLORS["white"])
    BORDER = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    CENTER = Alignment(horizontal="center", vertical="center")

    last_col = get_column_letter(len(COLS_CORR) + 1)

    TDB4.merge_cells(f"A1:{last_col}1")
    TDB4["A1"] = "Tableau de Bord 4 : Carte des Corrélations de Santé Mentale"
    TDB4["A1"].fill, TDB4["A1"].font = FILL_DARK, FONT_MAIN
    TDB4["A1"].alignment = CENTER
    TDB4.row_dimensions[1].height = 35

    start_row = 4
    for col_idx, var in enumerate(COLS_CORR):
        col_letter = get_column_letter(col_idx + 2)
        TDB4[f"{col_letter}{start_row}"] = var
        TDB4[f"{col_letter}{start_row}"].fill = FILL_HEAD
        TDB4[f"{col_letter}{start_row}"].font = FONT_HEAD
        TDB4[f"{col_letter}{start_row}"].alignment = CENTER
        TDB4[f"{col_letter}{start_row}"].border = BORDER

    for row_idx, var in enumerate(COLS_CORR):
        current_row = start_row + 1 + row_idx
        TDB4.row_dimensions[current_row].height = 20
        TDB4[f"A{current_row}"] = var
        TDB4[f"A{current_row}"].font = Font(bold=True, size=10)
        TDB4[f"A{current_row}"].border = BORDER

        for col_idx in range(len(COLS_CORR)):
            col_letter = get_column_letter(col_idx + 2)
            corr_row = 3 + row_idx
            corr_col = get_column_letter(2 + col_idx)
            TDB4[f"{col_letter}{current_row}"] = f"=Correlations!{corr_col}{corr_row}"
            TDB4[f"{col_letter}{current_row}"].number_format = "0.00"
            TDB4[f"{col_letter}{current_row}"].alignment = CENTER
            TDB4[f"{col_letter}{current_row}"].border = BORDER

    TDB4.column_dimensions["A"].width = 24
    for col_idx in range(len(COLS_CORR)):
        TDB4.column_dimensions[get_column_letter(col_idx + 2)].width = 22

    color_scale = ColorScaleRule(
        start_type="num",
        start_value=-1.0,
        start_color="FFF2F2",
        mid_type="num",
        mid_value=0.0,
        mid_color="FFFFFF",
        end_type="num",
        end_value=1.0,
        end_color="E6F2FF",
    )
    matrix_range = (
        f"B5:{get_column_letter(len(COLS_CORR) + 1)}{start_row + len(COLS_CORR)}"
    )
    TDB4.conditional_formatting.add(matrix_range, color_scale)

    wb.save(PATH_FILE)
    wb.close()
    print("TDB4 créée (heatmap corrélations)")


# ─── UTILITAIRE COMMUN ────────────────────────────────────────────────────────


def _apply_series_colors(chart, colors: list) -> None:
    """
    Applique une couleur de remplissage à chaque série d'un graphique BarChart.

    Pour chaque série, la couleur est appliquée à deux niveaux :
      1. La série entière (``graphicalProperties.solidFill``) → couleur par défaut
         de toutes les barres de la série.
      2. Chaque point de données individuel (``data_points``) → écrase la couleur
         par défaut série par série, ce qui est nécessaire pour que Excel
         l'interprète correctement quand les données sont dynamiques.

    Parameters
    ----------
    chart : BarChart
        Le graphique openpyxl à coloriser.
    colors : list of str
        Liste de couleurs hexadécimales sans '#' (ex. ["040459", "7D093F"]).
        La couleur à l'indice i est appliquée à la série i.
        Si ``colors`` est plus long que le nombre de séries, le surplus est ignoré.

    Examples
    --------
    >>> _apply_series_colors(chart, [COLORS["male"], COLORS["female"]])
    # Série 0 (Male)   → bleu très foncé  #040459
    # Série 1 (Female) → bordeaux          #7D093F
    """
    for i, color in enumerate(colors):
        if i < len(chart.series):
            chart.series[i].graphicalProperties.solidFill = color
            for pt in chart.series[i].data_points:
                pt.graphicalProperties.solidFill = color
