"""
sheets/dashboards.py
--------------------
Création des pages de tableau de bord : TDB1, TDB2, TDB3.
"""

from openpyxl import load_workbook

from config.settings import PATH_FILE
from projet_sante_mentale.utils.excel_helpers import setup_dashboard_page, setup_filter_row


TITLE = "Impact des réseaux sociaux sur la santé mentale"


def build_tdb1(len_dict: dict) -> None:
    """Crée la page TDB1 : titre + filtres."""
    wb = load_workbook(PATH_FILE)

    if "TDB1" in wb.sheetnames:
        del wb["TDB1"]
    ws = wb.create_sheet("TDB1", 0)

    setup_dashboard_page(ws, TITLE)
    setup_filter_row(ws, filter_row=3, len_dict=len_dict, tdb_prefix="TDB1")

    wb.save(PATH_FILE)
    wb.close()
    print("✅ TDB1 créée (titre + filtres)")


def build_tdb2(len_dict: dict) -> None:
    """Crée la page TDB2 : titre + filtres."""
    wb = load_workbook(PATH_FILE)

    if "TDB2" in wb.sheetnames:
        del wb["TDB2"]
    ws = wb.create_sheet("TDB2", 1)

    setup_dashboard_page(ws, TITLE)
    setup_filter_row(ws, filter_row=3, len_dict=len_dict, tdb_prefix="TDB2")

    wb.save(PATH_FILE)
    wb.close()
    print("✅ TDB2 créée (titre + filtres)")


def build_tdb3() -> None:
    """Crée la page TDB3 : titre uniquement (graphiques ajoutés par charts.py)."""
    wb = load_workbook(PATH_FILE)

    if "TDB3" in wb.sheetnames:
        del wb["TDB3"]
    ws = wb.create_sheet("TDB3", 2)

    setup_dashboard_page(ws, "Analyse des corrélations avancées (Santé Mentale)")

    wb.save(PATH_FILE)
    wb.close()
    print("✅ TDB3 créée (titre, graphiques à ajouter via charts)")
