"""
sheets/tcd2.py
--------------
Création de la feuille TCD2 : tableaux croisés pour TDB2.
  - Tableau 1 : Performance scolaire par Sexe et Temps d'écran
  - Tableau 2 : Dépressifs par Âge
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from config.settings import PATH_FILE
from utils.styles import (
    TITLE_FILL,
    TITLE_FONT_SM,
    HEADER_FILL,
    HEADER_FONT,
    ROW_HEADER_FILL,
    ROW_HEADER_FONT,
    BORDER,
    CENTER,
)
from utils.excel_helpers import write_section_title, write_col_headers


TEMPS_ECRAN = [("0-1h", 0, 1), ("1-2h", 1, 2), ("2-3h", 2, 3), ("3-4h", 3, 4)]
SEXES = ["Male", "Female"]


def build_tcd2_sheet(len_dict: dict) -> None:
    """
    Orchestre la création de la feuille 'TCD2' (Tableaux Croisés Dynamiques 2).

    Cette fonction initialise la feuille (en écrasant une éventuelle version précédente)
    et appelle les sous-fonctions chargées de construire les différents tableaux 
    de données nécessaires au Tableau de Bord 2 (TDB2).

    Parameters
    ----------
    len_dict : dict
        Dictionnaire contenant les dimensions (nombre de valeurs uniques) des variables 
        catégorielles (ex: 'len_age', 'len_gender'), utilisé pour dimensionner 
        les tableaux dynamiquement.

    Returns
    -------
    None
        Sauvegarde le fichier Excel après la création de la feuille.
    """
    wb = load_workbook(PATH_FILE)

    if "TCD2" in wb.sheetnames:
        del wb["TCD2"]
    ws = wb.create_sheet("TCD2")

    _write_perf_scolaire(ws)
    _write_depressifs_par_age(ws, len_dict)

    wb.save(PATH_FILE)
    wb.close()
    print("✅ Feuille TCD2 créée (perf scolaire + dépressifs par âge)")


# ─── TABLEAU 1 : Performance Scolaire ─────────────────────────────────────────
def _write_perf_scolaire(ws) -> None:
    """
    Génère le Tableau 1 : Performance Scolaire Moyenne par Sexe et Temps d'Écran.

    Construit un tableau croisé avec les sexes en lignes et les tranches de temps 
    d'écran en colonnes. Les valeurs sont calculées via la formule Excel `AVERAGEIFS` 
    et sont dynamiquement liées aux filtres du TDB2 (Âge, Plateforme, Interaction).

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        La feuille 'TCD2' en cours d'édition.

    Returns
    -------
    None
        Modifie la feuille de calcul en place.
    """
    start = 1

    ws[f"A{start}"] = "Performance Scolaire Moyenne par Sexe et Temps d'Écran"
    ws.merge_cells(f"A{start}:M{start}")
    ws[f"A{start}"].font = TITLE_FONT_SM
    ws[f"A{start}"].fill = TITLE_FILL
    ws[f"A{start}"].alignment = CENTER

    # En-têtes
    header_row = 2
    ws[f"A{header_row}"] = "Sexe"
    ws[f"A{header_row}"].fill = HEADER_FILL
    ws[f"A{header_row}"].border = BORDER

    for i, (label, _, __) in enumerate(TEMPS_ECRAN):
        col = get_column_letter(i + 2)
        ws[f"{col}{header_row}"] = label
        ws[f"{col}{header_row}"].fill = HEADER_FILL
        ws[f"{col}{header_row}"].border = BORDER
        ws[f"{col}{header_row}"].alignment = CENTER

    # Données
    for i, sexe in enumerate(SEXES):
        row = 3 + i
        ws[f"A{row}"] = sexe
        ws[f"A{row}"].fill = ROW_HEADER_FILL
        ws[f"A{row}"].font = ROW_HEADER_FONT
        ws[f"A{row}"].border = BORDER
        ws[f"A{row}"].alignment = CENTER

        for j, (_, min_v, max_v) in enumerate(TEMPS_ECRAN):
            col = get_column_letter(j + 2)
            formula = (
                f'=IFERROR(AVERAGEIFS(DATA!$G:$G, DATA!$B:$B, "{sexe.lower()}", '
                f'DATA!$F:$F, ">="&{min_v}, DATA!$F:$F, "<"&{max_v}, '
                f'DATA!$A:$A, IF(TDB2!$D$3="Tous", "<>", TDB2!$D$3), '
                f'DATA!$D:$D, IF(TDB2!$J$3="Tous", "<>", TDB2!$J$3), '
                f'DATA!$I:$I, IF(TDB2!$M$3="Tous", "<>", TDB2!$M$3)), 0)'
            )
            ws[f"{col}{row}"] = formula
            ws[f"{col}{row}"].number_format = "0.0"
            ws[f"{col}{row}"].border = BORDER
            ws[f"{col}{row}"].alignment = CENTER


# ─── TABLEAU 2 : Dépressifs par Âge ──────────────────────────────────────────
def _write_depressifs_par_age(ws, len_dict: dict) -> None:
    """
    Génère le Tableau 2 : Nombre d'individus dépressifs par tranche d'Âge.

    Construit un tableau affichant de manière conditionnelle les tranches d'âges 
    (selon le filtre sélectionné dans TDB2) et comptabilise le nombre d'individus 
    dépressifs via la formule Excel `COUNTIFS`. Les résultats respectent l'ensemble 
    des filtres appliqués sur le TDB2.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        La feuille 'TCD2' en cours d'édition.
    len_dict : dict
        Dictionnaire des dimensions, utilisé pour itérer dynamiquement sur le 
        nombre d'âges uniques présents dans les données.

    Returns
    -------
    None
        Modifie la feuille de calcul en place.
    """
    start = 10

    write_section_title(ws, start, "Nombre de dépressifs par Âge", f"A{start}:B{start}")

    header_row = start + 1
    write_col_headers(ws, header_row, ["Âge", "Nombre de dépressifs"], widths=[12, 18])

    for age_idx in range(1, len_dict["len_age"]):
        row_num = header_row + age_idx

        # Colonne A : Âge conditionnel
        ws[f"A{row_num}"] = (
            f'=IF(OR(TDB2!$D$3="Tous", IFERROR(INDEX(CALC!$C$2:$C$100,{age_idx}),)=TDB2!$D$3), '
            f'IFERROR(INDEX(CALC!$C$2:$C$100,{age_idx}),""), "")'
        )
        ws[f"A{row_num}"].fill = ROW_HEADER_FILL
        ws[f"A{row_num}"].font = ROW_HEADER_FONT
        ws[f"A{row_num}"].border = BORDER
        ws[f"A{row_num}"].alignment = CENTER

        # Colonne B : COUNTIFS dépressifs
        ws[f"B{row_num}"] = (
            f'=IF(OR(TDB2!$D$3="Tous", IFERROR(INDEX(CALC!$C$2:$C$100,{age_idx}),)=TDB2!$D$3), '
            f"IFERROR(COUNTIFS("
            f"DATA!$M$2:$M$1201, 1, "
            f'DATA!$A$2:$A$1201, IF(TDB2!$D$3="Tous", IFERROR(INDEX(CALC!$C$2:$C$100,{age_idx}),), TDB2!$D$3), '
            f'DATA!$B$2:$B$1201, IF(TDB2!$G$3="Tous", "<>", TDB2!$G$3), '
            f'DATA!$D$2:$D$1201, IF(TDB2!$J$3="Tous", "<>", TDB2!$J$3), '
            f'DATA!$I$2:$I$1201, IF(TDB2!$M$3="Tous", "<>", TDB2!$M$3)), 0), "")'
        )
        ws[f"B{row_num}"].border = BORDER
        ws[f"B{row_num}"].alignment = CENTER
        ws[f"B{row_num}"].number_format = "0"
