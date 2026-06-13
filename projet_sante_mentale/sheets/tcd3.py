"""
sheets/tcd3.py
--------------
Création de la feuille TCD3 : 4 tableaux pour les scatter plots de TDB3.
"""

from openpyxl import load_workbook

from config.settings import PATH_FILE
from utils.styles import HEADER_FILL, HEADER_FONT, BORDER, CENTER
from utils.excel_helpers import write_section_title, write_col_headers, write_row_header


SLEEP_VALS = [4.0, 4.5, 5.0, 5.5, 6.0, 6.5, 7.0, 7.5, 8.0, 8.5]
SCREEN_VALS = [0.5, 1.0, 1.5, 2.0, 2.5, 3.0, 3.5]
STRESS_VALS = list(range(1, 11))


def build_tcd3_sheet() -> None:
    """
    Orchestre la création de la feuille 'TCD3' (Tableaux Croisés Dynamiques 3).

    Cette fonction initialise la feuille (en l'écrasant si elle existe déjà) et 
    appelle successivement les sous-fonctions chargées de construire les 4 tableaux 
    de données. Ces tableaux alimenteront les graphiques en nuage de points (Scatter Plots) 
    du Tableau de Bord 3.

    Returns
    -------
    None
        Sauvegarde le fichier Excel après la génération.
    """
    wb = load_workbook(PATH_FILE)

    if "TCD3" in wb.sheetnames:
        del wb["TCD3"]
    ws = wb.create_sheet("TCD3")

    _write_sleep_vs_addiction(ws)  # Lignes  1-12
    _write_perf_vs_stress(ws)  # Lignes 15-26
    _write_anxiety_vs_screen(ws)  # Lignes 30-39
    _write_perf_vs_sleep_dep(ws)  # Lignes 42-54

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    wb.save(PATH_FILE)
    wb.close()
    print("Feuille TCD3 créée (4 tableaux scatter)")


# ─── Fonction générique d'écriture d'un tableau 3 colonnes ───────────────────
def _write_scatter_table(
    ws, title_row: int, title: str, headers: list, rows_data: list
) -> None:
    """
    Génère et formate un tableau générique à 3 colonnes destiné à un graphique de type Scatter.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        La feuille 'TCD3' en cours d'édition.
    title_row : int
        L'index de la ligne où insérer le titre du tableau.
    title : str
        Le titre textuel du tableau.
    headers : list of str
        Liste contenant les 3 libellés d'en-tête (ex: ["Axe X", "Série 1", "Série 2"]).
    rows_data : list of tuples
        Liste contenant les données de chaque ligne sous la forme :
        (valeur_x, formule_colonne_B, formule_colonne_C, format_x, format_y).

    Returns
    -------
    None
        Modifie la feuille de calcul en place.
    """
    write_section_title(ws, title_row, title, f"A{title_row}:C{title_row}")
    write_col_headers(ws, title_row + 1, headers, widths=[16, 18, 18])

    for idx, (x_val, f_b, f_c, fmt_x, fmt_y) in enumerate(rows_data):
        row = title_row + 2 + idx
        write_row_header(ws, row, 1, x_val)
        ws.cell(row=row, column=1).number_format = fmt_x

        for col, formula, fmt in [(2, f_b, fmt_y), (3, f_c, fmt_y)]:
            cell = ws.cell(row=row, column=col, value=formula)
            cell.number_format = fmt
            cell.border = BORDER
            cell.alignment = CENTER


# ─── TABLEAU 1 : Sommeil vs Addiction ────────────────────────────────────────
def _write_sleep_vs_addiction(ws) -> None:
    """
    Génère le Tableau 1 : Sommeil vs Addiction selon l'Interaction Sociale.

    Crée les données pour un nuage de points croisant les heures de sommeil (Axe X) 
    et le niveau d'addiction (Axe Y), scindé en deux séries (Interaction faible / élevée).

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        La feuille cible.
    """
    rows = []
    for v in SLEEP_VALS:
        rows.append(
            (
                v,
                f'=IFERROR(AVERAGEIFS(DATA!$L$2:$L$1201, DATA!$E$2:$E$1201, ">="&({v}-0.25), '
                f'DATA!$E$2:$E$1201, "<"&({v}+0.25), DATA!$I$2:$I$1201, "low"), 0)',
                f'=IFERROR(AVERAGEIFS(DATA!$L$2:$L$1201, DATA!$E$2:$E$1201, ">="&({v}-0.25), '
                f'DATA!$E$2:$E$1201, "<"&({v}+0.25), DATA!$I$2:$I$1201, "high"), 0)',
                "0.0",
                "0.00",
            )
        )
    _write_scatter_table(
        ws,
        1,
        "Sommeil vs Addiction par Interaction Sociale",
        ["Sommeil (h)", "Addiction (Low)", "Addiction (High)"],
        rows,
    )


# ─── TABLEAU 2 : Performance vs Stress ───────────────────────────────────────
def _write_perf_vs_stress(ws) -> None:
    """
    Génère le Tableau 2 : Performance Scolaire vs Stress selon le Genre.

    Crée les données pour un nuage de points croisant le niveau de stress (Axe X) 
    et la performance scolaire (Axe Y), scindé en deux séries (Male / Female).

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        La feuille cible.
    """
    rows = []
    for v in STRESS_VALS:
        rows.append(
            (
                v,
                f'=IFERROR(AVERAGEIFS(DATA!$G$2:$G$1201, DATA!$J$2:$J$1201, {v}, DATA!$B$2:$B$1201, "male"), 0)',
                f'=IFERROR(AVERAGEIFS(DATA!$G$2:$G$1201, DATA!$J$2:$J$1201, {v}, DATA!$B$2:$B$1201, "female"), 0)',
                "0",
                "0.00",
            )
        )
    _write_scatter_table(
        ws,
        15,
        "Performance vs Stress par Genre",
        ["Stress", "Performance (Male)", "Performance (Female)"],
        rows,
    )


# ─── TABLEAU 3 : Anxiété vs Temps d'écran ────────────────────────────────────
def _write_anxiety_vs_screen(ws) -> None:
    """
    Génère le Tableau 3 : Anxiété vs Temps d'écran selon l'Interaction Sociale.

    Crée les données pour un nuage de points croisant le temps d'écran (Axe X) 
    et l'anxiété (Axe Y), scindé en deux séries (Interaction faible / élevée).

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        La feuille cible.
    """
    rows = []
    for v in SCREEN_VALS:
        rows.append(
            (
                v,
                f'=IFERROR(AVERAGEIFS(DATA!$K$2:$K$1201, DATA!$F$2:$F$1201, ">="&({v}-0.25), '
                f'DATA!$F$2:$F$1201, "<"&({v}+0.25), DATA!$I$2:$I$1201, "low"), 0)',
                f'=IFERROR(AVERAGEIFS(DATA!$K$2:$K$1201, DATA!$F$2:$F$1201, ">="&({v}-0.25), '
                f'DATA!$F$2:$F$1201, "<"&({v}+0.25), DATA!$I$2:$I$1201, "high"), 0)',
                "0.0",
                "0.00",
            )
        )
    _write_scatter_table(
        ws,
        30,
        "Anxiété vs Temps d'Écran par Interaction",
        ["Temps d'écran (h)", "Anxiété (Low)", "Anxiété (High)"],
        rows,
    )


# ─── TABLEAU 4 : Performance vs Sommeil par Dépression ───────────────────────
def _write_perf_vs_sleep_dep(ws) -> None:
    """
    Génère le Tableau 4 : Performance Scolaire vs Sommeil selon l'État Dépressif.

    Crée les données pour un nuage de points croisant les heures de sommeil (Axe X) 
    et la performance scolaire (Axe Y), scindé en deux séries (Dépressif / Non dépressif).

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        La feuille cible.
    """
    rows = []
    for v in SLEEP_VALS:
        rows.append(
            (
                v,
                f'=IFERROR(AVERAGEIFS(DATA!$G$2:$G$1201, DATA!$E$2:$E$1201, ">="&({v}-0.25), '
                f'DATA!$E$2:$E$1201, "<"&({v}+0.25), DATA!$M$2:$M$1201, 0), 0)',
                f'=IFERROR(AVERAGEIFS(DATA!$G$2:$G$1201, DATA!$E$2:$E$1201, ">="&({v}-0.25), '
                f'DATA!$E$2:$E$1201, "<"&({v}+0.25), DATA!$M$2:$M$1201, 1), 0)',
                "0.0",
                "0.00",
            )
        )
    _write_scatter_table(
        ws,
        42,
        "Performance vs Sommeil par État Dépressif",
        ["Sommeil (h)", "Performance (No Dep)", "Performance (Dep)"],
        rows,
    )
