"""
sheets/tcd.py
-------------
Création de la feuille TCD : tableaux croisés + KPI pour TDB1.
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.formula import ArrayFormula

from config.settings import PATH_FILE, COLORS, FILTER_TDB1, FILTER_ARR_TDB1
from utils.styles import (
    TITLE_FILL,
    TITLE_FONT_SM,
    HEADER_FILL,
    HEADER_FONT,
    ROW_HEADER_FILL,
    ROW_HEADER_FONT,
    KPI_TITLE_FILL,
    KPI_TITLE_FONT,
    KPI_VALUE_FILL,
    KPI_VALUE_FONT,
    KPI_VALUE_FONT_LG,
    BORDER,
    CENTER,
)
from utils.excel_helpers import write_section_title, write_col_headers


def build_tcd_sheet(len_dict: dict) -> None:
    """
    Orchestre la création complète de la feuille 'TCD' (Tableaux Croisés Dynamiques).

    Cette fonction principale initialise la feuille, supprime l'ancienne version si elle existe,
    et appelle successivement les sous-fonctions pour construire les différents tableaux 
    de données, la section KPI et les zones tampons (helpers) nécessaires aux graphiques 
    des tableaux de bord.

    Parameters
    ----------
    len_dict : dict
        Dictionnaire contenant les dimensions (nombres de valeurs uniques) des différentes 
        variables catégorielles, utilisé pour dimensionner dynamiquement les tableaux.

    Returns
    -------
    None
        Sauvegarde les modifications directement dans le fichier Excel spécifié par PATH_FILE.
    """
    wb = load_workbook(PATH_FILE)

    if "TCD" in wb.sheetnames:
        del wb["TCD"]
    ws = wb.create_sheet("TCD")

    _write_tableau_depression(ws, len_dict)
    _write_tableau_addiction(ws, len_dict)
    _write_tableau_boxplot(ws, len_dict)
    _write_kpi_section(ws)
    _write_pie_helper_tdb1(ws)
    _write_helpers_tdb2(ws)

    wb.save(PATH_FILE)
    wb.close()
    print("✅ Feuille TCD créée (4 tableaux + KPI + helpers)")


# ─── TABLEAU 1 : Dépression par Genre et Âge ──────────────────────────────────
def _write_tableau_depression(ws, len_dict: dict) -> None:
    """
    Génère le Tableau 1 : Comptage des individus dépressifs par Genre (lignes) et par Âge (colonnes).

    Construit un tableau croisé dynamique via des formules Excel `COUNTIFS` 
    qui se mettent à jour automatiquement en fonction des filtres sélectionnés sur TDB1.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        La feuille 'TCD' en cours d'édition.
    len_dict : dict
        Dictionnaire des dimensions pour boucler sur le nombre d'âges et de genres.
    """
    write_section_title(ws, 1, "Dépression par Genre et Âge", "A1:J1")

    ws["A2"] = "Genre"
    ws["A2"].fill = HEADER_FILL
    ws["A2"].font = HEADER_FONT
    ws["A2"].border = BORDER
    ws["A2"].alignment = CENTER
    ws.column_dimensions["A"].width = 15

    for col_idx in range(1, len_dict["len_age"]):
        col_letter = get_column_letter(col_idx + 1)
        ws[f"{col_letter}2"] = (
            f'=IF(OR(TDB1!$D$3="Tous", IFERROR(INDEX(CALC!$C$2:$C$100,{col_idx}),)=TDB1!$D$3), '
            f'IFERROR(INDEX(CALC!$C$2:$C$100,{col_idx}),""), "")'
        )
        ws[f"{col_letter}2"].fill = HEADER_FILL
        ws[f"{col_letter}2"].font = HEADER_FONT
        ws[f"{col_letter}2"].border = BORDER
        ws[f"{col_letter}2"].alignment = CENTER
        ws.column_dimensions[col_letter].width = 12

    for row_idx in range(1, len_dict["len_gender"]):
        row_num = 2 + row_idx
        ws[f"A{row_num}"] = (
            f'=IF(OR(TDB1!$G$3="Tous", IFERROR(INDEX(CALC!$A$2:$A$100,{row_idx}),)=TDB1!$G$3), '
            f'IFERROR(INDEX(CALC!$A$2:$A$100,{row_idx}),""), "")'
        )
        ws[f"A{row_num}"].fill = ROW_HEADER_FILL
        ws[f"A{row_num}"].font = ROW_HEADER_FONT
        ws[f"A{row_num}"].border = BORDER
        ws[f"A{row_num}"].alignment = CENTER

        for col_idx in range(1, len_dict["len_age"]):
            col_letter = get_column_letter(col_idx + 1)
            formula = (
                f"=IF(AND("
                f'OR(TDB1!$D$3="Tous", IFERROR(INDEX(CALC!$C$2:$C$100,{col_idx}),)=TDB1!$D$3), '
                f'OR(TDB1!$G$3="Tous", IFERROR(INDEX(CALC!$A$2:$A$100,{row_idx}),)=TDB1!$G$3)), '
                f"IFERROR(COUNTIFS("
                f"DATA!$M:$M, 1, "
                f'DATA!$B:$B, IF(TDB1!$G$3="Tous", IFERROR(INDEX(CALC!$A$2:$A$100,{row_idx}),""), TDB1!$G$3), '
                f'DATA!$A:$A, IF(TDB1!$D$3="Tous", IFERROR(INDEX(CALC!$C$2:$C$100,{col_idx}),), TDB1!$D$3), '
                f'DATA!$D:$D, IF(TDB1!$J$3="Tous", "<>", TDB1!$J$3), '
                f'DATA!$I:$I, IF(TDB1!$M$3="Tous", "<>", TDB1!$M$3)), 0), "")'
            )
            ws[f"{col_letter}{row_num}"] = formula
            ws[f"{col_letter}{row_num}"].border = BORDER
            ws[f"{col_letter}{row_num}"].alignment = CENTER
            ws[f"{col_letter}{row_num}"].number_format = "0"


# ─── TABLEAU 2 : Addiction Moyenne par Genre et Âge ───────────────────────────
def _write_tableau_addiction(ws, len_dict: dict) -> None:
    """
    Génère le Tableau 2 : Calcul du niveau d'addiction moyen par Genre et Âge.

    Fonctionne de manière similaire à _write_tableau_depression, mais utilise la 
    formule Excel `AVERAGEIFS` pour faire ressortir une moyenne conditionnelle.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        La feuille 'TCD' en cours d'édition.
    len_dict : dict
        Dictionnaire des dimensions pour construire la grille dynamiquement.
    """
    start = 10
    write_section_title(
        ws, start, "Addiction Moyenne par Genre et Âge", f"A{start}:J{start}"
    )

    header_row = start + 1
    ws[f"A{header_row}"] = "Genre"
    ws[f"A{header_row}"].fill = HEADER_FILL
    ws[f"A{header_row}"].font = HEADER_FONT
    ws[f"A{header_row}"].border = BORDER
    ws[f"A{header_row}"].alignment = CENTER

    for col_idx in range(1, len_dict["len_age"]):
        col_letter = get_column_letter(col_idx + 1)
        ws[f"{col_letter}{header_row}"] = (
            f'=IF(OR(TDB1!$D$3="Tous", IFERROR(INDEX(CALC!$C$2:$C$100,{col_idx}),)=TDB1!$D$3), '
            f'IFERROR(INDEX(CALC!$C$2:$C$100,{col_idx}),""), "")'
        )
        ws[f"{col_letter}{header_row}"].fill = HEADER_FILL
        ws[f"{col_letter}{header_row}"].font = HEADER_FONT
        ws[f"{col_letter}{header_row}"].border = BORDER
        ws[f"{col_letter}{header_row}"].alignment = CENTER

    for row_idx in range(1, len_dict["len_gender"]):
        row_num = header_row + row_idx
        ws[f"A{row_num}"] = (
            f'=IF(OR(TDB1!$G$3="Tous", IFERROR(INDEX(CALC!$A$2:$A$100,{row_idx}),)=TDB1!$G$3), '
            f'IFERROR(INDEX(CALC!$A$2:$A$100,{row_idx}),""), "")'
        )
        ws[f"A{row_num}"].fill = ROW_HEADER_FILL
        ws[f"A{row_num}"].font = ROW_HEADER_FONT
        ws[f"A{row_num}"].border = BORDER
        ws[f"A{row_num}"].alignment = CENTER

        for col_idx in range(1, len_dict["len_age"]):
            col_letter = get_column_letter(col_idx + 1)
            formula = (
                f"=IF(AND("
                f'OR(TDB1!$D$3="Tous", IFERROR(INDEX(CALC!$C$2:$C$100,{col_idx}),)=TDB1!$D$3), '
                f'OR(TDB1!$G$3="Tous", IFERROR(INDEX(CALC!$A$2:$A$100,{row_idx}),)=TDB1!$G$3)), '
                f"IFERROR(AVERAGEIFS("
                f"DATA!$L:$L, "
                f'DATA!$B:$B, IF(TDB1!$G$3="Tous", IFERROR(INDEX(CALC!$A$2:$A$100,{row_idx}),""), TDB1!$G$3), '
                f'DATA!$A:$A, IF(TDB1!$D$3="Tous", IFERROR(INDEX(CALC!$C$2:$C$100,{col_idx}),), TDB1!$D$3), '
                f'DATA!$D:$D, IF(TDB1!$J$3="Tous", "<>", TDB1!$J$3), '
                f'DATA!$I:$I, IF(TDB1!$M$3="Tous", "<>", TDB1!$M$3)), 0), "")'
            )
            ws[f"{col_letter}{row_num}"] = formula
            ws[f"{col_letter}{row_num}"].border = BORDER
            ws[f"{col_letter}{row_num}"].alignment = CENTER
            ws[f"{col_letter}{row_num}"].number_format = "0.00"


# ─── TABLEAU 3 : Statistiques Boxplot par Âge ─────────────────────────────────
def _write_tableau_boxplot(ws, len_dict: dict) -> None:
    """
    Génère le Tableau 3 : Statistiques de base de l'addiction pour chaque tranche d'Âge.

    Calcule et affiche le minimum, le 1er quartile, la médiane, le 3ème quartile, le maximum,
    la moyenne et l'écart-type. Les formules matricielles (ArrayFormula) sont massivement 
    utilisées pour filtrer la plage de calcul en respectant le contexte interactif.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        La feuille 'TCD' en cours d'édition.
    len_dict : dict
        Dictionnaire des dimensions, utilisé ici principalement pour le nombre d'âges.
    """
    start = 20
    write_section_title(
        ws, start, "Statistiques Addiction par Âge (pour Boxplot)", f"A{start}:H{start}"
    )

    headers = ["Âge", "Min", "Q1", "Médiane", "Q3", "Max", "Moyenne", "Écart-type"]
    write_col_headers(ws, start + 1, headers, widths=[12] * 8)

    num_ages = len_dict["len_age"] - 1
    for row_idx in range(1, num_ages + 1):
        row_num = start + 1 + row_idx

        ws[f"A{row_num}"] = f'=IFERROR(INDEX(CALC!$C$2:$C$1000,{row_idx}),"")'
        ws[f"A{row_num}"].fill = ROW_HEADER_FILL
        ws[f"A{row_num}"].font = ROW_HEADER_FONT
        ws[f"A{row_num}"].border = BORDER
        ws[f"A{row_num}"].alignment = CENTER

        stats = [
            (
                "B",
                f"=MIN(IF((DATA!$A$2:$A$1201=$A{row_num}){FILTER_ARR_TDB1}, DATA!$L$2:$L$1201))",
            ),
            (
                "C",
                f"=QUARTILE(IF((DATA!$A$2:$A$1201=$A{row_num}){FILTER_ARR_TDB1}, DATA!$L$2:$L$1201), 1)",
            ),
            (
                "D",
                f"=MEDIAN(IF((DATA!$A$2:$A$1201=$A{row_num}){FILTER_ARR_TDB1}, DATA!$L$2:$L$1201))",
            ),
            (
                "E",
                f"=QUARTILE(IF((DATA!$A$2:$A$1201=$A{row_num}){FILTER_ARR_TDB1}, DATA!$L$2:$L$1201), 3)",
            ),
            (
                "F",
                f"=MAX(IF((DATA!$A$2:$A$1201=$A{row_num}){FILTER_ARR_TDB1}, DATA!$L$2:$L$1201))",
            ),
            (
                "H",
                f"=STDEV(IF((DATA!$A$2:$A$1201=$A{row_num}){FILTER_ARR_TDB1}, DATA!$L$2:$L$1201))",
            ),
        ]
        for col_l, formula in stats:
            ws[f"{col_l}{row_num}"] = ArrayFormula(f"{col_l}{row_num}", formula)

        ws[f"G{row_num}"] = (
            f"=IFERROR(AVERAGEIFS(DATA!$L$2:$L$1201, DATA!$A$2:$A$1201, $A{row_num}, "
            f"DATA!$M$2:$M$1201, 1{FILTER_TDB1}), NA())"
        )

        for col_idx in range(2, 9):
            c = get_column_letter(col_idx)
            ws[f"{c}{row_num}"].border = BORDER
            ws[f"{c}{row_num}"].alignment = CENTER
            ws[f"{c}{row_num}"].number_format = "0.00"


# ─── KPI SECTION ──────────────────────────────────────────────────────────────
def _write_kpi_section(ws) -> None:
    """
    Construit la section des Indicateurs Clés de Performance (KPI) ciblés sur les individus dépressifs.

    Génère des blocs visuellement mis en avant contenant :
    - Le réseau social le plus utilisé et le total des dépressifs.
    - La performance scolaire moyenne, avec ses bornes minimales et maximales.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        La feuille 'TCD' cible.
    """
    kpi_start = 30

    write_section_title(
        ws,
        kpi_start,
        "Indicateurs Clés (KPI) - Personnes Dépressives",
        f"A{kpi_start}:H{kpi_start}",
    )

    kpi1_row = kpi_start + 2
    _write_kpi_block(
        ws,
        kpi1_row,
        "A",
        "C",
        title="Réseau le plus utilisé",
        value_formula=(
            f"=INDEX(DATA!$D$2:$D$1201, MATCH(MAX(COUNTIFS("
            f"DATA!$D$2:$D$1201, DATA!$D$2:$D$1201, DATA!$M$2:$M$1201, 1{FILTER_TDB1})), "
            f"COUNTIFS(DATA!$D$2:$D$1201, DATA!$D$2:$D$1201, "
            f"DATA!$M$2:$M$1201, 1{FILTER_TDB1}), 0))"
        ),
        use_array=True,
        num_fmt="@",
    )

    _write_kpi_count(
        ws,
        kpi1_row + 2,
        "A",
        "B",
        label="Nombre de dépressifs",
        formula=f"=IFERROR(COUNTIFS(DATA!$M$2:$M$1201, 1{FILTER_TDB1}), 0)",
    )

    _write_kpi_block(
        ws,
        kpi1_row,
        "D",
        "F",
        title="Performance scolaire moyenne",
        value_formula=(
            f"=IFERROR(AVERAGEIFS(DATA!$G$2:$G$1201, DATA!$M$2:$M$1201, 1{FILTER_TDB1}), NA())"
        ),
        num_fmt="0.00",
    )

    min_max_row = kpi1_row + 2
    ws[f"D{min_max_row}"] = "Min / Max"
    ws[f"D{min_max_row}"].font = Font(bold=True, size=10)
    ws[f"D{min_max_row}"].border = BORDER
    ws[f"D{min_max_row}"].alignment = Alignment(horizontal="left", vertical="center")

    ws[f"E{min_max_row}"] = ArrayFormula(
        f"E{min_max_row}",
        f"=MIN(IF((DATA!$M$2:$M$1201=1){FILTER_ARR_TDB1}, DATA!$G$2:$G$1201))",
    )
    ws[f"E{min_max_row}"].border = BORDER
    ws[f"E{min_max_row}"].alignment = CENTER
    ws[f"E{min_max_row}"].number_format = "0.00"

    ws[f"F{min_max_row}"] = ArrayFormula(
        f"F{min_max_row}",
        f"=MAX(IF((DATA!$M$2:$M$1201=1){FILTER_ARR_TDB1}, DATA!$G$2:$G$1201))",
    )
    ws[f"F{min_max_row}"].border = BORDER
    ws[f"F{min_max_row}"].alignment = CENTER
    ws[f"F{min_max_row}"].number_format = "0.00"

    for col in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 18


def _write_kpi_block(
    ws,
    row: int,
    col_start: str,
    col_end: str,
    title: str,
    value_formula: str,
    use_array: bool = False,
    num_fmt: str = "0.00",
) -> None:
    """
    Fonction interne factorisant la création stylisée d'un bloc de KPI principal.

    Dessine un rectangle comprenant une barre de titre en haut et une large zone de valeur
    centrée et formatée en dessous. Gère optionnellement la création d'une formule matricielle.

    Parameters
    ----------
    ws : Worksheet
        Feuille de travail cible.
    row : int
        Index de la ligne où débute le bloc.
    col_start : str
        Lettre de la colonne de départ pour la fusion (ex: 'A').
    col_end : str
        Lettre de la colonne de fin pour la fusion (ex: 'C').
    title : str
        Intitulé textuel affiché dans l'en-tête du KPI.
    value_formula : str
        La formule Excel stricte à évaluer pour obtenir la valeur de l'indicateur.
    use_array : bool, optional
        Si True, enveloppe la formule dans un objet openpyxl ArrayFormula.
    num_fmt : str, optional
        Format numérique à appliquer à la cellule de résultat (ex: '@' pour string, '0.00' pour décimal).
    """
    ws[f"{col_start}{row}"] = title
    ws[f"{col_start}{row}"].fill = KPI_TITLE_FILL
    ws[f"{col_start}{row}"].font = KPI_TITLE_FONT
    ws[f"{col_start}{row}"].border = BORDER
    ws[f"{col_start}{row}"].alignment = CENTER
    ws.merge_cells(f"{col_start}{row}:{col_end}{row}")
    ws.row_dimensions[row].height = 20

    value_row = row + 1
    if use_array:
        ws[f"{col_start}{value_row}"] = ArrayFormula(
            f"{col_start}{value_row}", value_formula
        )
    else:
        ws[f"{col_start}{value_row}"] = value_formula
    ws[f"{col_start}{value_row}"].fill = KPI_VALUE_FONT_LG and KPI_TITLE_FILL
    ws[f"{col_start}{value_row}"].font = KPI_VALUE_FONT_LG
    ws[f"{col_start}{value_row}"].border = BORDER
    ws[f"{col_start}{value_row}"].alignment = CENTER
    ws[f"{col_start}{value_row}"].number_format = num_fmt
    ws.merge_cells(f"{col_start}{value_row}:{col_end}{value_row}")
    ws.row_dimensions[value_row].height = 35


def _write_kpi_count(
    ws, row: int, label_col: str, val_col: str, label: str, formula: str
) -> None:
    """
    Fonction interne générant une ligne d'information secondaire (label / valeur) sous un KPI.

    Parameters
    ----------
    ws : Worksheet
        Feuille de travail cible.
    row : int
        Index de la ligne pour insérer la sous-information.
    label_col : str
        Colonne accueillant le libellé textuel.
    val_col : str
        Colonne accueillant le résultat de la formule.
    label : str
        Le libellé (ex: 'Nombre de dépressifs').
    formula : str
        La formule Excel renvoyant le compte / résultat.
    """
    ws[f"{label_col}{row}"] = label
    ws[f"{label_col}{row}"].font = Font(bold=True, size=10)
    ws[f"{label_col}{row}"].border = BORDER
    ws[f"{label_col}{row}"].alignment = Alignment(horizontal="left", vertical="center")

    ws[f"{val_col}{row}"] = formula
    ws[f"{val_col}{row}"].font = Font(bold=True, size=11)
    ws[f"{val_col}{row}"].border = BORDER
    ws[f"{val_col}{row}"].alignment = CENTER
    ws[f"{val_col}{row}"].number_format = "0"


# ─── HELPERS PieChart TDB1 ────────────────────────────────────────────────────


def _write_pie_helper_tdb1(ws) -> None:
    """
    Dresse un mini-tableau tampon au bas de la feuille 'TCD' (ligne 120) 
    servant de source de données spécifiquement au graphique en secteurs (PieChart) du TDB1.

    Calcule la répartition de la dépression selon le genre tout en respectant l'ensemble des filtres du TDB1.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Feuille de travail cible.
    """
    row = 120
    ws[f"A{row}"] = "Total Dépression Genre (TDB1)"
    ws[f"A{row}"].font = Font(bold=True, size=9)

    ws[f"A{row + 1}"] = "Genre"
    ws[f"B{row + 1}"] = "Total"

    for i, (genre, label) in enumerate(
        [("male", "Male"), ("female", "Female")], start=2
    ):
        ws[f"A{row + i}"] = label
        ws[f"B{row + i}"] = (
            f'=IF(OR(TDB1!$G$3="Tous", TDB1!$G$3="{label}"), '
            f'IFERROR(COUNTIFS(DATA!$M$2:$M$1201, 1, DATA!$B$2:$B$1201, "{genre}", '
            f'DATA!$D$2:$D$1201, IF(TDB1!$J$3="Tous", "<>", TDB1!$J$3), '
            f'DATA!$I$2:$I$1201, IF(TDB1!$M$3="Tous", "<>", TDB1!$M$3), '
            f'DATA!$A$2:$A$1201, IF(TDB1!$D$3="Tous", "<>", TDB1!$D$3)), 0), 0)'
        )
        ws[f"B{row + i}"].number_format = "0"


# ─── HELPERS TDB2 ─────────────────────────────────────────────────────────────
def _write_helpers_tdb2(ws) -> None:
    """
    Prépare deux tableaux tampons supplémentaires (lignes 100 et 105) dédiés 
    à l'alimentation des graphiques du Tableau de Bord 2 (TDB2).

    Le premier tableau compte le nombre de dépressifs. 
    Le second évalue le niveau moyen d'addiction. 
    Ces calculs écoutent dynamiquement les sélecteurs de filtres positionnés sur la feuille TDB2.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Feuille de travail cible.
    """
    for helper_row, col_data, label_title in [
        (100, "M", "Dépressifs"),
        (105, "L", "Addiction Moyenne"),
    ]:
        ws[f"A{helper_row}"] = "Genre"
        ws[f"B{helper_row}"] = label_title
        ws[f"A{helper_row}"].font = Font(bold=True)
        ws[f"B{helper_row}"].font = Font(bold=True)

        for offset, (genre, label) in enumerate(
            [("male", "Male"), ("female", "Female")], start=1
        ):
            ws[f"A{helper_row + offset}"] = label

            if col_data == "M":  # COUNTIFS
                formula = (
                    f'=IF(OR(TDB2!$G$3="Tous", TDB2!$G$3="{label}"), '
                    f'IFERROR(COUNTIFS(DATA!$M$2:$M$1201, 1, DATA!$B$2:$B$1201, "{genre}", '
                    f'DATA!$D$2:$D$1201, IF(TDB2!$J$3="Tous", "<>", TDB2!$J$3), '
                    f'DATA!$I$2:$I$1201, IF(TDB2!$M$3="Tous", "<>", TDB2!$M$3), '
                    f'DATA!$A$2:$A$1201, IF(TDB2!$D$3="Tous", "<>", TDB2!$D$3)), 0), 0)'
                )
                ws[f"B{helper_row + offset}"].number_format = "0"
            else:  # AVERAGEIFS
                formula = (
                    f'=IF(OR(TDB2!$G$3="Tous", TDB2!$G$3="{label}"), '
                    f'IFERROR(AVERAGEIFS(DATA!$L$2:$L$1201, DATA!$B$2:$B$1201, "{genre}", '
                    f'DATA!$D$2:$D$1201, IF(TDB2!$J$3="Tous", "<>", TDB2!$J$3), '
                    f'DATA!$I$2:$I$1201, IF(TDB2!$M$3="Tous", "<>", TDB2!$M$3), '
                    f'DATA!$A$2:$A$1201, IF(TDB2!$D$3="Tous", "<>", TDB2!$D$3)), 0), 0)'
                )
                ws[f"B{helper_row + offset}"].number_format = "0.0"

            ws[f"B{helper_row + offset}"] = formula
