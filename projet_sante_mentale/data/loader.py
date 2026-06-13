"""
data/loader.py
--------------
Chargement, nettoyage et préparation du DataFrame.
"""

import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table, TableStyleInfo
from config.settings import (
    DATA_URL,
    PATH_FILE,
    TEMPLATE_DIR,
    PATH_FILE,
    COLS_FOR_LEN,
    COLS_CORR,
)


# ─── CHARGEMENT CSV ───────────────────────────────────────────────────────────


def load_data() -> pd.DataFrame:
    """Charge le CSV et enrichit les colonnes encodées."""
    df = pd.read_csv(DATA_URL)

    # Encodages numériques pour la matrice de corrélation
    df["gender_num"] = df["gender"].map({"male": 0, "female": 1})
    df["social_num"] = df["social_interaction_level"].map(
        {"low": 0, "medium": 1, "high": 2}
    )

    print(f"✅ Données chargées : {len(df)} lignes, {len(df.columns)} colonnes")
    print(f"   Valeurs manquantes : {df.isnull().sum().sum()}")
    return df


# ─── CALCUL DE LEN_DICT ───────────────────────────────────────────────────────


def build_len_dict(df: pd.DataFrame) -> dict:
    """
    Retourne un dictionnaire {len_<col>: nb_valeurs_uniques + 1}
    pour les colonnes définies dans COLS_FOR_LEN.
    """
    len_dict = {f"len_{col}": len(df[col].unique()) + 1 for col in COLS_FOR_LEN}
    print(f"✅ len_dict construit : {len_dict}")
    return len_dict


# ─── ÉCRITURE DU FICHIER EXCEL ────────────────────────────────────────────────


def init_workbook(df: pd.DataFrame) -> None:
    """
    Crée ou recharge le fichier Excel et y injecte la feuille DATA.
    """
    os.makedirs(os.path.dirname(PATH_FILE), exist_ok=True)

    if not os.path.exists(PATH_FILE):
        wb = Workbook()
        wb.save(PATH_FILE)

    with pd.ExcelWriter(PATH_FILE, mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name="DATA", index=False)

    # Supprimer la feuille vide par défaut
    wb = load_workbook(PATH_FILE)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(PATH_FILE)
    wb.close()

    print(f"✅ Feuille DATA écrite → {PATH_FILE}")


# ─── FEUILLE CALC ─────────────────────────────────────────────────────────────


def build_calc_sheet(len_dict: dict) -> None:
    """
    Crée ou recrée la feuille CALC avec les valeurs distinctes
    (ArrayFormula UNIQUE) pour : Genres, Âges, Plateformes, Interactions.
    """
    wb = load_workbook(PATH_FILE)

    if "CALC" not in wb.sheetnames:
        ws_calc = wb.create_sheet("CALC")
    else:
        ws_calc = wb["CALC"]

    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    _unique_col(ws_calc, "A", "DATA!B:B", len_dict["len_gender"], "tblGenres", style)
    _unique_col(ws_calc, "C", "DATA!A:A", len_dict["len_age"], "tblAges", style)
    _unique_col(
        ws_calc,
        "E",
        "DATA!D:D",
        len_dict["len_platform_usage"],
        "tblPlateformes",
        style,
    )
    _unique_col(
        ws_calc,
        "G",
        "DATA!I:I",
        len_dict["len_social_interaction_level"],
        "tblInteractions",
        style,
    )

    wb.save(PATH_FILE)
    wb.close()
    print("✅ Feuille CALC créée avec les valeurs distinctes")


def _unique_col(ws, col: str, source: str, length: int, table_name: str, style) -> None:
    """Écrit une ArrayFormula UNIQUE dans une colonne et crée un Table."""
    ws[f"{col}1"] = ArrayFormula(f"{col}1:{col}{length}", f"=_xlfn.UNIQUE({source})")
    tbl = Table(displayName=table_name, ref=f"{col}1:{col}{length}")
    tbl.tableStyleInfo = style
    tbl.hasHeader = False
    ws.add_table(tbl)


# ─── MATRICE DE CORRÉLATION ───────────────────────────────────────────────────
def build_correlations_sheet(df: pd.DataFrame) -> None:
    """Calcule la matrice de corrélation et l'écrit dans la feuille Correlations."""
    from utils.excel_helpers import write_table

    corr = df[COLS_CORR].corr().round(2)
    corr_reset = corr.reset_index()
    corr_reset.columns = ["Variable"] + COLS_CORR

    wb = load_workbook(PATH_FILE)
    if "Correlations" not in wb.sheetnames:
        ws_corr = wb.create_sheet("Correlations")
    else:
        ws_corr = wb["Correlations"]

    write_table(
        ws_corr, corr_reset, start_row=1, title="Matrice de corrélation", space=0
    )

    wb.save(PATH_FILE)
    wb.close()
    print("Feuille Correlations créée")
