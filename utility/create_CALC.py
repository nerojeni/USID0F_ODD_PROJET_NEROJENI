import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table, TableStyleInfo
from utility.excel_utils import get_or_create_sheet


def create_calc_sheet(df: pd.DataFrame, file_path: str) -> dict:
    """
    Crée la feuille 'CALC' dans le fichier Excel contenant les valeurs uniques
    des colonnes spécifiées, utilisées pour les filtres dynamiques.

    Parameters
    ----------
    df : pd.DataFrame
        Le DataFrame contenant les données nettoyées.
    file_path : str
        Le chemin vers le fichier Excel.

    Returns
    -------
    dict
        Un dictionnaire contenant les longueurs (nombre de valeurs uniques + 1)
        pour chaque colonne traitée. (ex: {'len_age': 8, 'len_gender': 3})
    """
    print("⚙️ Création de la feuille de calculs 'CALC'...")

    # 1. Calcul des longueurs dynamiques (len_dict)
    cols_to_calculate = ["age", "gender", "platform_usage", "social_interaction_level"]
    len_dict = {}
    for col in cols_to_calculate:
        # +1 pour laisser de la place à l'en-tête ou à la formule
        len_dict[f"len_{col}"] = len(df[col].unique()) + 1

    print(f"   ℹ️ Dimensions trouvées : {len_dict}")

    # 2. Chargement du fichier Excel
    wb = load_workbook(file_path)

    ws_calc = get_or_create_sheet(wb, "CALC", reset=True)

    # Style pour les tableaux
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    # ========== GENRES (Colonne A) ==========
    formula_gender = "=_xlfn.UNIQUE(DATA!B:B)"
    ws_calc["A1"] = ArrayFormula(f"A1:A{len_dict['len_gender']}", formula_gender)
    table_gender = Table(displayName="tblGenres", ref=f"A1:A{len_dict['len_gender']}")
    table_gender.tableStyleInfo = style
    table_gender.hasHeader = False
    ws_calc.add_table(table_gender)

    # ========== AGES (Colonne C) ==========
    formula_age = "=_xlfn.UNIQUE(DATA!A:A)"
    ws_calc["C1"] = ArrayFormula(f"C1:C{len_dict['len_age']}", formula_age)
    table_age = Table(displayName="tblAges", ref=f"C1:C{len_dict['len_age']}")
    table_age.tableStyleInfo = style
    table_age.hasHeader = False
    ws_calc.add_table(table_age)

    # ========== PLATEFORMES (Colonne E) ==========
    formula_platform = "=_xlfn.UNIQUE(DATA!D:D)"
    ws_calc["E1"] = ArrayFormula(
        f"E1:E{len_dict['len_platform_usage']}", formula_platform
    )
    table_platform = Table(
        displayName="tblPlateformes", ref=f"E1:E{len_dict['len_platform_usage']}"
    )
    table_platform.tableStyleInfo = style
    table_platform.hasHeader = False
    ws_calc.add_table(table_platform)

    # ========== INTERACTIONS SOCIALES (Colonne G) ==========
    formula_interaction = "=_xlfn.UNIQUE(DATA!I:I)"
    ws_calc["G1"] = ArrayFormula(
        f"G1:G{len_dict['len_social_interaction_level']}", formula_interaction
    )
    table_interaction = Table(
        displayName="tblInteractions",
        ref=f"G1:G{len_dict['len_social_interaction_level']}",
    )
    table_interaction.tableStyleInfo = style
    table_interaction.hasHeader = False
    ws_calc.add_table(table_interaction)

    # Sauvegarde
    wb.save(file_path)
    print(f"✅ Feuille 'CALC' créée avec succès !")
    print(f"📋 Feuilles actuelles : {wb.sheetnames}\n")
    wb.close()

    return len_dict
