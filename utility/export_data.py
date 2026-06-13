import os
import pandas as pd
from openpyxl import load_workbook, Workbook


def save_data_to_excel(df: pd.DataFrame, file_path: str) -> None:
    """
    Exporte un DataFrame pandas vers la feuille 'DATA' d'un fichier Excel.
    Crée le fichier Excel de base s'il n'existe pas encore.

    Parameters
    ----------
    df : pd.DataFrame
        Le DataFrame contenant les données nettoyées à exporter.
    file_path : str
        Le chemin complet vers le fichier Excel de destination.

    Returns
    -------
    None
        Modifie ou crée le fichier Excel et affiche un rapport dans la console.
    """
    print(f"Sauvegarde des données dans la feuille 'DATA'...")

    # 1. S'assurer que le dossier de destination existe
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    # 2. Vérifier si le fichier existe et est valide, sinon créer un fichier vide
    if not os.path.exists(file_path):
        wb = Workbook()
        wb.save(file_path)
        wb.close()
    else:
        try:
            # Test d'ouverture pour s'assurer que le fichier n'est pas corrompu
            wb = load_workbook(file_path)
            wb.close()
        except Exception:
            print(
                "Fichier existant corrompu ou invalide. Recréation d'un fichier propre."
            )
            wb = Workbook()
            wb.save(file_path)
            wb.close()

    # 3. Ajouter ou remplacer la feuille 'DATA' via pandas
    try:
        with pd.ExcelWriter(
            file_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            df.to_excel(writer, sheet_name="DATA", index=False)

        # 4. Vérification finale
        wb_check = load_workbook(file_path)
        sheet_names = wb_check.sheetnames
        wb_check.close()

        # Nettoyage optionnel : supprimer la feuille par défaut "Sheet" si elle est vide
        if "Sheet" in sheet_names and len(sheet_names) > 1:
            wb_clean = load_workbook(file_path)
            del wb_clean["Sheet"]
            wb_clean.save(file_path)
            sheet_names = wb_clean.sheetnames
            wb_clean.close()

        print("Données sauvegardées avec succès !")
        print(f"Feuilles actuellement présentes : {sheet_names}\n")

    except Exception as e:
        print(f"Erreur lors de l'écriture dans le fichier Excel : {e}")
