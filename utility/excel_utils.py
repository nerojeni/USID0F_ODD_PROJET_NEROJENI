from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.series import SeriesLabel
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from utility.config import COLOR_BLUE_DARK, COLOR_TEAL

# =================================================================
# DÉFINITION DES STYLES GLOBAUX
# =================================================================
title_fill = PatternFill(
    start_color=COLOR_BLUE_DARK, end_color=COLOR_BLUE_DARK, fill_type="solid"
)
title_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(
    start_color=COLOR_TEAL, end_color=COLOR_TEAL, fill_type="solid"
)
header_font = Font(color="FFFFFF", bold=True, size=10)
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
center_align = Alignment(horizontal="center", vertical="center")


def style_header(ws, row: int, cols: list = ["A", "B", "C"]) -> None:
    """
    Applique le style d'en-tête standard aux cellules d'une ligne spécifique.

    Cette fonction applique une couleur de fond (Teal), une police blanche et grasse,
    des bordures fines ainsi qu'un alignement centré sur une liste de colonnes définies.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        La feuille de calcul cible.
    row : int
        L'index de la ligne contenant les en-têtes à styliser (ex: 2).
    cols : list, optional
        La liste des lettres de colonnes à modifier. Par défaut, ["A", "B", "C"].

    Returns
    -------
    None
        Modifie la feuille de calcul en place.
    """
    for col in cols:
        ws[f"{col}{row}"].fill = header_fill
        ws[f"{col}{row}"].font = header_font
        ws[f"{col}{row}"].border = border
        ws[f"{col}{row}"].alignment = center_align


def style_title(ws, row: int, text: str, span: str = "A:C") -> None:
    """
    Crée et stylise une barre de titre principale fusionnée sur une ligne.

    Fusionne les cellules spécifiées par le paramètre `span`, y insère le texte du titre,
    applique un fond bleu foncé, une police blanche grasse, un alignement centré
    et définit une hauteur de ligne fixe de 25 pixels.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        La feuille de calcul cible.
    row : int
        L'index de la ligne où le titre doit être créé (ex: 1).
    text : str
        Le texte à afficher dans le titre.
    span : str, optional
        La plage de colonnes à fusionner sous la forme 'Début:Fin'. Par défaut, "A:C".

    Returns
    -------
    None
        Modifie la feuille de calcul en place.
    """
    start, end = span.split(":")
    ws[f"{start}{row}"] = text
    ws[f"{start}{row}"].font = title_font
    ws[f"{start}{row}"].fill = title_fill
    ws.merge_cells(f"{start}{row}:{end}{row}")
    ws.row_dimensions[row].height = 25


def add_filter(
    worksheet,
    start_col: str,
    filter_row: int,
    title_text: str,
    data_source_col: str,
    len_data: int,
    helper_col: str,
    default_value: str = "Tous",
    title_color: str = "FF003D6B",
    value_color: str = "FFFFFF",
) -> None:
    """
    Crée un filtre horizontal déroulant interactif avec un titre et sa valeur côte à côte.

    Le filtre génère automatiquement une colonne technique masquée (helper_col) pour inclure
    l'option universelle "Tous" sans altérer la feuille CALC d'origine. Il applique une règle
    de validation de données (DataValidation) sous forme de liste déroulante sur la cellule de valeur.

    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        La feuille de travail (Tableau de bord) où le filtre doit être inséré.
    start_col : str
        La lettre de la colonne où sera écrit le titre du filtre (ex: 'C').
        La cellule de sélection se placera automatiquement sur la colonne suivante (ex: 'D').
    filter_row : int
        L'index de la ligne où positionner le bloc de filtre (ex: 3).
    title_text : str
        Le libellé affiché sur le bouton de titre (ex: 'Âge', 'Genre').
    data_source_col : str
        La lettre de la colonne source contenant les données uniques dans la feuille 'CALC'.
    len_data : int
        Le nombre de lignes de données uniques à lire dans la feuille 'CALC'.
    helper_col : str
        La lettre de la colonne technique à utiliser pour stocker la liste temporaire (ex: 'AE').
    default_value : str, optional
        La valeur texte initialement sélectionnée à l'ouverture. Par défaut, 'Tous'.
    title_color : str, optional
        Code couleur hexadécimal (ARGB) pour le fond du titre. Par défaut, "FF003D6B" (bleu foncé).
    value_color : str, optional
        Code couleur hexadécimal (ARGB) pour le fond de la cellule de sélection. Par défaut, "FFFFFF" (blanc).

    Returns
    -------
    None
        Modifie la feuille de calcul en place et configure la validation de données.
    """
    # Colonne de la valeur (juste après le titre)
    value_col = chr(ord(start_col) + 1)

    # Remplissage pour titre et valeur
    current_title_fill = PatternFill(
        start_color=title_color, end_color=title_color, fill_type="solid"
    )
    current_value_fill = PatternFill(
        start_color=value_color, end_color=value_color, fill_type="solid"
    )

    alignment = Alignment(horizontal="center", vertical="center")
    cell_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    title_font_filter = Font(bold=True, color="FFFFFF", size=10)
    value_font_filter = Font(bold=True, color="000000", size=10)

    # ========== TITRE DU FILTRE ==========
    title_cell = worksheet[f"{start_col}{filter_row}"]
    title_cell.value = title_text
    title_cell.alignment = alignment
    title_cell.fill = current_title_fill
    title_cell.border = cell_border
    title_cell.font = title_font_filter
    worksheet.column_dimensions[start_col].width = 12

    # ========== CELLULE DE VALEUR ==========
    value_cell = worksheet[f"{value_col}{filter_row}"]
    value_cell.value = default_value
    value_cell.alignment = alignment
    value_cell.fill = current_value_fill
    value_cell.border = cell_border
    value_cell.font = value_font_filter
    worksheet.column_dimensions[value_col].width = 12

    # ========== COLONNE HELPER (CACHÉE) ==========
    worksheet[f"{helper_col}1"] = "Tous"

    for i in range(2, len_data + 1):
        worksheet[f"{helper_col}{i}"] = f"=CALC!{data_source_col}{i}"

    worksheet.column_dimensions[helper_col].hidden = True

    # ========== VALIDATION DE DONNÉES ==========
    formula = f"=${helper_col}$1:${helper_col}${len_data}"

    dv = DataValidation(type="list", formula1=formula, allow_blank=False)
    dv.error = "Sélectionnez une valeur valide"
    dv.errorTitle = "Entrée invalide"
    dv.prompt = f"Sélectionnez un {title_text.lower()}"
    dv.promptTitle = "Filtrer"

    worksheet.add_data_validation(dv)
    dv.add(f"{value_col}{filter_row}")

    print(
        f"✅ Filtre '{title_text}' créé en {start_col}{filter_row}:{value_col}{filter_row}"
    )


def write_table(
    ws,
    df: pd.DataFrame,
    start_row: int,
    title: str = None,
    space: int = 3,
    padding: int = 2,
) -> int:
    """
    Écrit un DataFrame dans une feuille Excel OpenPyXL à partir d'une ligne donnée.

    La fonction ajoute éventuellement un titre, écrit les en-têtes de colonnes
    puis les données du DataFrame. Elle retourne ensuite la première ligne
    disponible pour écrire un nouveau tableau en laissant un nombre de lignes
    vides configurable.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Feuille OpenPyXL cible.
    df : pd.DataFrame
        DataFrame à écrire.
    start_row : int
        Ligne de départ.
    title : str, optional
        Titre du tableau. Par défaut None.
    space : int, optional
        Nombre de lignes vides à laisser après le tableau. Par défaut 3.
    padding : int, optional
        Espace supplémentaire pour la largeur des colonnes. Par défaut 2.

    Returns
    -------
    int
        Numéro de la prochaine ligne disponible pour le tableau suivant.
    """
    # Écriture du titre s'il existe
    if title:
        ws.cell(row=start_row, column=1, value=title)
        start_row += 1

    # En-têtes + ajustement automatique de la largeur des colonnes
    for col_idx, header in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=col_idx, value=header)

        col_letter = get_column_letter(col_idx)
        width = len(str(header)) + padding

        # On conserve la plus grande largeur si la colonne a déjà été élargie avant
        current_width = ws.column_dimensions[col_letter].width
        if current_width is None or width > current_width:
            ws.column_dimensions[col_letter].width = width

    # Données (On utilise enumerate sur df.values pour éviter les problèmes d'index pandas)
    for r_idx, row_data in enumerate(df.values, start=1):
        for c_idx, value in enumerate(row_data, start=1):
            ws.cell(row=start_row + r_idx, column=c_idx, value=value)

    # Retourne la ligne de départ théorique du tableau suivant
    return start_row + len(df) + space + 1


def get_or_create_sheet(wb: Workbook, sheet_name: str, reset: bool = False) -> Worksheet:
    """
    Récupère une feuille de calcul existante ou la crée si elle n'existe pas.

    Parameters
    ----------
    wb : openpyxl.workbook.Workbook
        Le classeur Excel en cours de manipulation.
    sheet_name : str
        Le nom de la feuille à récupérer ou à créer.
    reset : bool, optional
        Si True, supprime la feuille existante pour en recréer une totalement vierge. 
        Par défaut False.

    Returns
    -------
    openpyxl.worksheet.worksheet.Worksheet
        L'objet Worksheet correspondant à la feuille demandée.
    """
    if sheet_name in wb.sheetnames:
        if reset:
            del wb[sheet_name]
            ws = wb.create_sheet(sheet_name)
            print(f"🔄 Feuille '{sheet_name}' réinitialisée.")
        else:
            ws = wb[sheet_name]
            print(f"📂 Feuille '{sheet_name}' récupérée (déjà existante).")
    else:
        ws = wb.create_sheet(sheet_name)
        print(f"✨ Feuille '{sheet_name}' créée.")
        
    return ws